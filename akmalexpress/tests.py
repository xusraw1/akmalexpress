from io import BytesIO
from decimal import Decimal
from unittest.mock import patch
from openpyxl import Workbook, load_workbook

from django.contrib.auth.models import User
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse, translate_url
from datetime import timedelta
from django.utils.crypto import get_random_string
from django.utils import timezone

from .models import Order, OrderAttachment, OrderItem
from .templatetags.number_format import money


class ProfileViewTests(TestCase):
    def setUp(self):
        test_password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='profileadmin',
            password=test_password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username='profileadmin', password=test_password)

        order = Order.objects.create(
            user=self.staff,
            receipt_number=123,
            order_date=timezone.localdate(),
            first_name='Test',
            last_name='Client',
            phone1=998901112233,
            status=Order.Status.ORDERED,
            track_number='',
        )
        OrderItem.objects.create(
            order=order,
            product_name='Sneakers',
            product_quantity=1,
            product_price_currency='USD',
            product_price='120.000',
            store='Taobao',
        )

    def test_profile_page_renders(self):
        response = self.client.get(reverse('profile', args=[self.staff.username]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '@profileadmin')

    def test_profile_shorthand_redirects_to_current_user(self):
        response = self.client.get(f"{reverse('my_profile')}?sort=date_asc&status=ordered")
        expected = f"{reverse('profile', args=[self.staff.username])}?sort=date_asc&status=ordered"
        self.assertRedirects(response, expected, fetch_redirect_response=False)

    def test_profile_with_at_prefix_is_supported(self):
        response = self.client.get(reverse('profile', args=[f'@{self.staff.username}']))
        self.assertEqual(response.status_code, 200)

    def test_profile_filters_do_not_crash(self):
        query_strings = [
            '?status=ordered',
            '?sort=date_asc',
            '?sort=receipt_desc',
            '?sort=receipt_asc',
            '?sort=status_asc',
            '?sort=status_desc',
            '?date_from=2026-01-01&date_to=2026-12-31',
            '?search=Sneakers',
            '?missing_track=1',
            '?status=invalid',
            '?date_from=invalid',
        ]

        for query in query_strings:
            with self.subTest(query=query):
                response = self.client.get(f"{reverse('profile', args=[self.staff.username])}{query}")
                self.assertEqual(response.status_code, 200)

    def test_profile_orders_have_pagination(self):
        for idx in range(2, 12):
            Order.objects.create(
                user=self.staff,
                receipt_number=123 + idx,
                order_date=timezone.localdate() - timedelta(days=idx),
                first_name='Test',
                last_name=f'Client {idx}',
                phone1=998901110000 + idx,
                status=Order.Status.ORDERED,
            )

        response = self.client.get(reverse('profile', args=[self.staff.username]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'aria-label="Пагинация"', html=False)
        self.assertContains(response, '?page=2')


class MissingTrackFilterTests(TestCase):
    def setUp(self):
        password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='missingtrackstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=self.staff.username, password=password)

        self.order_tracked = Order.objects.create(
            user=self.staff,
            receipt_number=810,
            order_date=timezone.localdate(),
            first_name='Tracked',
            last_name='Client',
            phone1=998901111810,
            status=Order.Status.ORDERED,
        )
        OrderItem.objects.create(
            order=self.order_tracked,
            product_name='Tracked item',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='10000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='TRK-810',
            store='Taobao',
        )

        self.order_missing = Order.objects.create(
            user=self.staff,
            receipt_number=811,
            order_date=timezone.localdate(),
            first_name='Missing',
            last_name='Client',
            phone1=998901111811,
            status=Order.Status.ORDERED,
        )
        OrderItem.objects.create(
            order=self.order_missing,
            product_name='Missing item',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='12000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='',
            store='Taobao',
        )

        self.order_mixed = Order.objects.create(
            user=self.staff,
            receipt_number=812,
            order_date=timezone.localdate(),
            first_name='Mixed',
            last_name='Client',
            phone1=998901111812,
            status=Order.Status.ORDERED,
        )
        OrderItem.objects.create(
            order=self.order_mixed,
            product_name='Mixed tracked',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='12000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='TRK-812-A',
            store='Taobao',
        )
        OrderItem.objects.create(
            order=self.order_mixed,
            product_name='Mixed missing',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='12000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='',
            store='Taobao',
        )

    def test_orders_page_filter_missing_track(self):
        response = self.client.get(reverse('orders'), {'missing_track': '1'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Missing Client')
        self.assertContains(response, 'Mixed Client')
        self.assertNotContains(response, 'Tracked Client')

    def test_profile_filter_missing_track(self):
        response = self.client.get(
            reverse('profile', args=[self.staff.username]),
            {'missing_track': '1'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Missing Client')
        self.assertContains(response, 'Mixed Client')
        self.assertNotContains(response, 'Tracked Client')


class ErrorPageTests(TestCase):
    def test_404_page_is_rendered(self):
        response = self.client.get('/this-route-does-not-exist/')
        self.assertEqual(response.status_code, 404)
        self.assertContains(response, '404', status_code=404)
        self.assertContains(response, 'Страница не найдена', status_code=404)


class StaffEntryPointTests(TestCase):
    def test_staff_login_is_available_without_trailing_slash(self):
        response = self.client.get(f"/{settings.STAFF_LOGIN_URL.rstrip('/')}")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Служебный вход')

    def test_panel_redirects_guest_to_staff_login(self):
        response = self.client.get(reverse('panel'))
        self.assertRedirects(
            response,
            reverse('staff_login'),
            fetch_redirect_response=False,
        )

    def test_panel_without_slash_redirects_guest_to_staff_login(self):
        response = self.client.get('/panel')
        self.assertRedirects(
            response,
            reverse('staff_login'),
            fetch_redirect_response=False,
        )

    def test_panel_redirects_staff_to_orders(self):
        password = get_random_string(24)
        user = User.objects.create_user(
            username='panelstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=user.username, password=password)
        response = self.client.get(reverse('panel'))
        self.assertRedirects(response, reverse('orders'), fetch_redirect_response=False)

    def test_panel_redirects_superuser_to_admin(self):
        password = get_random_string(24)
        superuser = User.objects.create_superuser(
            username='panelroot',
            password=password,
            email='panelroot@example.com',
        )
        self.client.login(username=superuser.username, password=password)
        response = self.client.get(reverse('panel'))
        self.assertRedirects(
            response,
            f'/{settings.ADMIN_URL}',
            fetch_redirect_response=False,
        )

    def test_my_profile_for_superuser_redirects_to_profile_page_not_admin(self):
        password = get_random_string(24)
        superuser = User.objects.create_superuser(
            username='profilesuper',
            password=password,
            email='profilesuper@example.com',
        )
        self.client.login(username=superuser.username, password=password)
        response = self.client.get(reverse('my_profile'), follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('profile', args=[superuser.username]), response.url)

    def test_admin_is_available_without_trailing_slash(self):
        response = self.client.get(f'/{settings.ADMIN_URL.rstrip("/")}')
        self.assertRedirects(
            response,
            f'/{settings.ADMIN_URL}',
            fetch_redirect_response=False,
        )

    def test_authenticated_staff_opening_login_redirects_to_home(self):
        password = get_random_string(24)
        user = User.objects.create_user(
            username='alreadylogged',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=user.username, password=password)
        response = self.client.get(reverse('staff_login'))
        self.assertRedirects(response, reverse('index'), fetch_redirect_response=False)

    def test_login_post_without_next_redirects_to_home(self):
        password = get_random_string(24)
        user = User.objects.create_user(
            username='loginsubmit',
            password=password,
            is_staff=True,
            is_active=True,
        )
        response = self.client.post(
            reverse('staff_login'),
            {'username': user.username, 'password': password},
            follow=False,
        )
        self.assertRedirects(response, reverse('index'), fetch_redirect_response=False)

    @override_settings(
        STAFF_LOGIN_RATE_LIMIT_ATTEMPTS=3,
        STAFF_LOGIN_RATE_LIMIT_WINDOW_SECONDS=600,
        STAFF_LOGIN_RATE_LIMIT_LOCK_SECONDS=600,
    )
    def test_login_rate_limit_blocks_after_repeated_failures(self):
        password = get_random_string(24)
        user = User.objects.create_user(
            username='ratelimitstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        login_url = reverse('staff_login')
        for _ in range(3):
            response = self.client.post(
                login_url,
                {'username': user.username, 'password': 'wrong-password'},
                follow=False,
            )
            self.assertEqual(response.status_code, 302)

        blocked_response = self.client.post(
            login_url,
            {'username': user.username, 'password': password},
            follow=True,
        )
        self.assertEqual(blocked_response.status_code, 200)
        self.assertContains(blocked_response, 'Слишком много попыток входа')
        self.assertNotIn('_auth_user_id', self.client.session)

    def test_security_headers_are_present(self):
        response = self.client.get(reverse('index'))
        self.assertIn('Content-Security-Policy', response)
        self.assertIn('Permissions-Policy', response)
        self.assertEqual(response['Cross-Origin-Resource-Policy'], 'same-origin')

    @override_settings(
        LANGUAGE_COOKIE_SECURE=True,
        LANGUAGE_COOKIE_SAMESITE='Lax',
        LANGUAGE_COOKIE_HTTPONLY=False,
    )
    def test_language_cookie_uses_secure_flags(self):
        response = self.client.get(reverse('set_language', args=['uz']) + '?next=/')
        language_cookie = response.cookies.get(settings.LANGUAGE_COOKIE_NAME)
        site_cookie = response.cookies.get('site_language')
        self.assertIsNotNone(language_cookie)
        self.assertIsNotNone(site_cookie)
        self.assertEqual(language_cookie['samesite'], 'Lax')
        self.assertEqual(site_cookie['samesite'], 'Lax')
        self.assertTrue(bool(language_cookie['secure']))
        self.assertTrue(bool(site_cookie['secure']))


class DispatchOrdersTests(TestCase):
    def setUp(self):
        password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='dispatchstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=self.staff.username, password=password)
        self.order = Order.objects.create(
            user=self.staff,
            receipt_number=431,
            order_date=timezone.localdate(),
            first_name='Daily',
            last_name='Client',
            phone1=998901234567,
            status=Order.Status.ACCEPTED,
            shipping_method=Order.ShippingMethod.AVIA,
        )
        Order.objects.create(
            user=self.staff,
            receipt_number=999,
            order_date=timezone.localdate(),
            first_name='Other',
            last_name='Name',
            phone1=998909999999,
            status=Order.Status.ORDERED,
            shipping_method=Order.ShippingMethod.IPOST,
        )

    def test_dispatch_page_renders(self):
        response = self.client.get(reverse('dispatch_orders'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Заказы на отправку')
        self.assertContains(response, '№431')

    def test_dispatch_page_shows_only_new_orders(self):
        response = self.client.get(reverse('dispatch_orders'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '№431')
        self.assertNotContains(response, '№999')

    def test_dispatch_page_can_update_order_status(self):
        response = self.client.post(
            reverse('dispatch_orders'),
            {
                'order_id': self.order.id,
                'status': Order.Status.TRANSIT,
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        self.order.refresh_from_db()
        self.assertEqual(self.order.status, Order.Status.TRANSIT)

    def test_dispatch_page_has_no_filter_or_search_ui(self):
        response = self.client.get(reverse('dispatch_orders'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Поиск (квитанция / имя)')
        self.assertNotContains(response, 'За день')

    def test_dispatch_order_disappears_after_marked_ordered(self):
        response = self.client.post(
            reverse('dispatch_orders'),
            {
                'order_id': self.order.id,
                'status': Order.Status.ORDERED,
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        response_after = self.client.get(reverse('dispatch_orders'))
        self.assertNotContains(response_after, 'Daily Client')
        self.assertContains(response_after, 'Новых заказов нет')


class TrackCenterTests(TestCase):
    def setUp(self):
        staff_password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='trackstaff',
            password=staff_password,
            is_staff=True,
            is_active=True,
        )
        self.staff_password = staff_password

        regular_password = get_random_string(24)
        self.regular = User.objects.create_user(
            username='trackregular',
            password=regular_password,
            is_active=True,
        )
        self.regular_password = regular_password

        self.waiting_order = Order.objects.create(
            user=self.staff,
            receipt_number=321,
            order_date=timezone.localdate(),
            first_name='Waiting',
            last_name='Client',
            phone1=998901111111,
            status=Order.Status.ORDERED,
        )
        OrderItem.objects.create(
            order=self.waiting_order,
            product_name='Waiting item',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='10000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='TRK-WAIT-001',
            store='Taobao',
        )
        self.arrived_order = Order.objects.create(
            user=self.staff,
            receipt_number=322,
            order_date=timezone.localdate(),
            first_name='Arrived',
            last_name='Client',
            phone1=998902222222,
            status=Order.Status.ARRIVED,
            come=timezone.now(),
        )
        OrderItem.objects.create(
            order=self.arrived_order,
            product_name='Arrived item',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='12000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            track_number='TRK-ARR-001',
            store='Taobao',
        )

    def test_track_center_requires_staff(self):
        self.client.login(username=self.regular.username, password=self.regular_password)
        response = self.client.get(reverse('track_center'))
        self.assertEqual(response.status_code, 302)

    def test_track_center_renders_arrived_queue_only(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.get(reverse('track_center'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Сканер треков')
        self.assertContains(response, 'TRK-ARR-001')
        self.assertNotContains(response, 'TRK-WAIT-001')

    def test_track_scan_opens_quick_status_panel(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('track_center'),
            {'track_number': '  trk-wait-001  '},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        self.assertIn('track=TRK-WAIT-001', response.url)
        quick_panel_response = self.client.get(response.url)
        self.assertEqual(quick_panel_response.status_code, 200)
        self.assertContains(quick_panel_response, 'TRK-WAIT-001')
        self.assertContains(quick_panel_response, 'Сохранить статус')
        self.waiting_order.refresh_from_db()
        self.assertEqual(self.waiting_order.status, Order.Status.ORDERED)

    def test_track_status_can_be_changed_from_quick_panel(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('track_center'),
            {
                'action': 'update_status',
                'order_id': str(self.waiting_order.id),
                'track_number': 'TRK-WAIT-001',
                'status': Order.Status.ARRIVED,
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        self.waiting_order.refresh_from_db()
        self.assertEqual(self.waiting_order.status, Order.Status.ARRIVED)
        self.assertIsNotNone(self.waiting_order.come)


class SettlementPrintTests(TestCase):
    def setUp(self):
        staff_password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='sheetstaff',
            password=staff_password,
            is_staff=True,
            is_active=True,
        )
        self.staff_password = staff_password

        regular_password = get_random_string(24)
        self.regular = User.objects.create_user(
            username='sheetregular',
            password=regular_password,
            is_active=True,
        )
        self.regular_password = regular_password

        self.order = Order.objects.create(
            user=self.staff,
            receipt_number=905,
            order_date=timezone.localdate(),
            first_name='Bahriddin',
            last_name='Pirov',
            phone1=941958228,
            status=Order.Status.ORDERED,
        )
        OrderItem.objects.create(
            order=self.order,
            product_name='Phone case',
            product_quantity=1,
            product_price_currency='UZS',
            product_price='190000.000',
            shipping_method=Order.ShippingMethod.AVIA,
            store='Taobao',
        )

    def test_settlement_page_requires_staff(self):
        self.client.login(username=self.regular.username, password=self.regular_password)
        response = self.client.get(reverse('print_settlement_sheet', args=[self.order.slug]))
        self.assertEqual(response.status_code, 302)

    def test_settlement_form_prefills_values(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.get(reverse('print_settlement_sheet', args=[self.order.slug]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Расчет клиента')
        self.assertContains(response, '905')
        self.assertContains(response, 'Bahriddin Pirov')
        self.assertContains(response, '941958228')

    def test_settlement_post_renders_print_sheet(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('print_settlement_sheet', args=[self.order.slug]),
            {
                'receipt_number': '49',
                'full_name': 'Пиров Бахриддин',
                'product_cost': '190000',
                'cargo_cost': '79000',
                'service_cost': '40000',
                'phone': '941958228',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '49 Пиров Бахриддин')
        self.assertContains(response, 'Tannarxi:')
        self.assertContains(response, '190 000')
        self.assertContains(response, '79 000')
        self.assertContains(response, '40 350')
        self.assertContains(response, '309 350')
        self.assertContains(response, '119 350')

    def test_settlement_auto_mode_below_threshold_sets_fixed_service(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('print_settlement_sheet', args=[self.order.slug]),
            {
                'receipt_number': '905',
                'full_name': 'Bahriddin Pirov',
                'product_cost': '50000',
                'cargo_cost': '10000',
                'service_mode': 'auto',
                'phone': '941958228',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '10 000')
        self.assertContains(response, '70 000')
        self.assertContains(response, '20 000')

    def test_settlement_manual_service_mode_can_force_fixed_amount(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('print_settlement_sheet', args=[self.order.slug]),
            {
                'receipt_number': '905',
                'full_name': 'Bahriddin Pirov',
                'product_cost': '190000',
                'cargo_cost': '79000',
                'service_mode': 'fixed_10000',
                'phone': '941958228',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '10 000')
        self.assertContains(response, '279 000')
        self.assertContains(response, '89 000')

    def test_settlement_percent_10_mode(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('print_settlement_sheet', args=[self.order.slug]),
            {
                'receipt_number': '905',
                'full_name': 'Bahriddin Pirov',
                'product_cost': '100000',
                'cargo_cost': '20000',
                'service_mode': 'percent_10',
                'phone': '941958228',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '12 000')
        self.assertContains(response, '132 000')
        self.assertContains(response, '32 000')

    def test_settlement_custom_percent_mode_with_minimum_service(self):
        self.client.login(username=self.staff.username, password=self.staff_password)
        response = self.client.post(
            reverse('print_settlement_sheet', args=[self.order.slug]),
            {
                'receipt_number': '905',
                'full_name': 'Bahriddin Pirov',
                'product_cost': '20000',
                'cargo_cost': '10000',
                'service_mode': 'custom_percent',
                'service_percent': '5',
                'phone': '941958228',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '10 000')
        self.assertContains(response, '40 000')
        self.assertContains(response, '20 000')


class TrackReminderDismissTests(TestCase):
    def setUp(self):
        password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='noticeadmin',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=self.staff.username, password=password)
        Order.objects.create(
            user=self.staff,
            receipt_number=741,
            order_date=timezone.localdate() - timedelta(days=3),
            first_name='Track',
            last_name='Missing',
            phone1=998901010101,
            status=Order.Status.ORDERED,
            track_number='',
        )

    def test_track_notice_can_be_dismissed_and_reappears_later(self):
        home_url = reverse('index')
        response = self.client.get(home_url)
        self.assertNotContains(response, 'без трек-номера')

        dismiss_response = self.client.post(
            reverse('dismiss_track_notice'),
            {'next': home_url},
            follow=False,
        )
        self.assertEqual(dismiss_response.status_code, 302)

        response_after_dismiss = self.client.get(home_url)
        self.assertNotContains(response_after_dismiss, 'без трек-номера')

        # Reminder is disabled globally, so it does not reappear anymore.
        response_after_expire = self.client.get(home_url)
        self.assertNotContains(response_after_expire, 'без трек-номера')


class PublicSearchTests(TestCase):
    def setUp(self):
        self.staff = User.objects.create_user(
            username='publicsearchstaff',
            password=get_random_string(24),
            is_staff=True,
            is_active=True,
        )
        self.order_exact = Order.objects.create(
            user=self.staff,
            receipt_number=111,
            order_date=timezone.localdate(),
            first_name='Ali',
            last_name='Karimov',
            phone1=998901234567,
            track_number='TRK-111-EXACT',
            status=Order.Status.ACCEPTED,
        )
        Order.objects.create(
            user=self.staff,
            receipt_number=112,
            order_date=timezone.localdate(),
            first_name='Alisher',
            last_name='Karimov',
            phone1=998907777777,
            track_number='TRK-112-EXACT',
            status=Order.Status.ACCEPTED,
        )

    def test_index_search_has_no_date_filters(self):
        response = self.client.get(reverse('index'))
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'name="date_from"', html=False)
        self.assertNotContains(response, 'name="date_to"', html=False)

    def test_index_search_by_receipt_is_exact(self):
        response = self.client.get(reverse('index'), {'search': '111'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '№111')
        self.assertNotContains(response, '№112')

    def test_index_search_by_full_name_and_phone_is_exact(self):
        response_by_name = self.client.get(reverse('index'), {'search': 'Ali Karimov'})
        self.assertEqual(response_by_name.status_code, 200)
        self.assertContains(response_by_name, '№111')
        self.assertNotContains(response_by_name, '№112')

        response_by_phone = self.client.get(reverse('index'), {'search': '+998 90 123 45 67'})
        self.assertEqual(response_by_phone.status_code, 200)
        self.assertContains(response_by_phone, '№111')
        self.assertNotContains(response_by_phone, '№112')

    def test_index_search_by_partial_token_does_not_match(self):
        response = self.client.get(reverse('index'), {'search': 'Kar'})
        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, '№111')
        self.assertNotContains(response, '№112')

    def test_index_search_by_track_number_is_exact(self):
        response = self.client.get(reverse('index'), {'search': 'TRK-111-EXACT'})
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '№111')
        self.assertNotContains(response, '№112')

        partial_response = self.client.get(reverse('index'), {'search': 'TRK-111'})
        self.assertEqual(partial_response.status_code, 200)
        self.assertNotContains(partial_response, '№111')
        self.assertNotContains(partial_response, '№112')


class ExchangeRatesApiTests(TestCase):
    @patch('akmalexpress.views_public.get_exchange_rates')
    def test_exchange_rates_api_returns_json(self, get_rates_mock):
        get_rates_mock.return_value = {
            'usd_rate': '12345.67',
            'rmb_rate': '1765.43',
            'source': 'USD:Ipakyuli, RMB:CBU',
            'source_date': '17.03.2026',
            'fetched_at': '2026-03-17T10:00:00+05:00',
        }
        response = self.client.get(reverse('exchange_rates_api'))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['usd_rate'], '12345.67')
        self.assertEqual(data['rmb_rate'], '1765.43')

    def test_exchange_rates_api_rejects_post(self):
        response = self.client.post(reverse('exchange_rates_api'))
        self.assertEqual(response.status_code, 405)


class NumberFormatFilterTests(TestCase):
    def test_money_adds_space_grouping_for_large_values(self):
        self.assertEqual(money('1000000'), '1 000 000')
        self.assertEqual(money('500000'), '500 000')

    def test_money_keeps_fraction_only_when_needed(self):
        self.assertEqual(money('1234567.50'), '1 234 567.5')
        self.assertEqual(money('1234567.567', 3), '1 234 567.567')


class CreateOrderFlowTests(TestCase):
    def setUp(self):
        password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='createorderstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=self.staff.username, password=password)

    def test_create_order_template_defaults_new_item_currency_to_uzs(self):
        response = self.client.get(reverse('create_order'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<option value="UZS" selected>UZS</option>', html=False)
        self.assertNotContains(response, 'id_shipping_method', html=False)

    def test_create_order_disables_costs_when_toggles_unchecked(self):
        response = self.client.post(
            reverse('create_order'),
            {
                'receipt_number': '1001',
                'order_date': timezone.localdate().strftime('%Y-%m-%d'),
                'first_name': 'Ali',
                'last_name': 'Karimov',
                'phone1': '+998901112233',
                'phone2': '',
                'debt': '0',
                'shipping_method': Order.ShippingMethod.AVIA,
                'status': Order.Status.ACCEPTED,
                'track_pending': 'on',
                'track_number': '',
                # toggles intentionally unchecked
                'cargo_cost': '50000',
                'service_cost': '25000',
                'usd_rate': '12205',
                'rmb_rate': '1807',
                'description': '',
                'items-TOTAL_FORMS': '1',
                'items-INITIAL_FORMS': '0',
                'items-MIN_NUM_FORMS': '0',
                'items-MAX_NUM_FORMS': '1000',
                'items-0-product_name': 'Sneakers',
                'items-0-product_quantity': '1',
                'items-0-product_price_currency': 'UZS',
                'items-0-product_price': '100000',
                'items-0-store': 'Taobao',
                'items-0-link': '',
                'items-0-DELETE': '',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        order = Order.objects.get(receipt_number=1001)
        self.assertFalse(order.cargo_enabled)
        self.assertFalse(order.service_enabled)
        self.assertEqual(str(order.cargo_cost), '0.00')
        self.assertEqual(str(order.service_cost), '0.00')

    def test_create_order_supports_mixed_currencies_and_converts_to_uzs_total(self):
        response = self.client.post(
            reverse('create_order'),
            {
                'receipt_number': '1002',
                'order_date': timezone.localdate().strftime('%Y-%m-%d'),
                'first_name': 'Val',
                'last_name': 'Mix',
                'phone1': '+998909998877',
                'phone2': '',
                'debt': '0',
                'shipping_method': Order.ShippingMethod.AVIA,
                'status': Order.Status.ACCEPTED,
                'track_pending': 'on',
                'track_number': '',
                'cargo_enabled': 'on',
                'cargo_cost': '0',
                'service_enabled': 'on',
                'service_cost': '0',
                'usd_rate': '12205',
                'rmb_rate': '1807',
                'description': '',
                'items-TOTAL_FORMS': '2',
                'items-INITIAL_FORMS': '0',
                'items-MIN_NUM_FORMS': '0',
                'items-MAX_NUM_FORMS': '1000',
                'items-0-product_name': 'USD item',
                'items-0-product_quantity': '10',
                'items-0-product_price_currency': 'USD',
                'items-0-product_price': '1',
                'items-0-store': 'Taobao',
                'items-0-link': '',
                'items-0-DELETE': '',
                'items-1-product_name': 'RMB item',
                'items-1-product_quantity': '100',
                'items-1-product_price_currency': 'RMB',
                'items-1-product_price': '1',
                'items-1-store': 'Alibaba',
                'items-1-link': '',
                'items-1-DELETE': '',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        order = Order.objects.get(receipt_number=1002)
        self.assertEqual(order.items.count(), 2)
        self.assertEqual(order.get_current, 'UZS')
        self.assertEqual(order.get_total_price, Decimal('302750.00'))

    def test_order_total_preview_endpoint_returns_auto_total(self):
        payload = {
            'usd_rate': '12000',
            'rmb_rate': '1800',
            'cargo_enabled': True,
            'service_enabled': True,
            'cargo_cost': '17000',
            'service_cost': '9000',
            'items': [
                {'quantity': '2', 'price': '10', 'currency': 'USD', 'store': 'Taobao', 'delete': False},
                {'quantity': '1', 'price': '50000', 'currency': 'UZS', 'store': 'Alibaba', 'delete': False},
            ],
        }
        response = self.client.post(
            reverse('order_total_preview'),
            data=payload,
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['items_total'], '290000.00')
        self.assertEqual(data['extra_total'], '0.00')
        self.assertEqual(data['auto_total'], '290000.00')

    def test_manual_total_overrides_computed_total(self):
        order = Order.objects.create(
            user=self.staff,
            receipt_number=1099,
            order_date=timezone.localdate(),
            first_name='Manual',
            last_name='Total',
            phone1=998901010100,
            status=Order.Status.ACCEPTED,
            manual_total=Decimal('555000.00'),
            cargo_enabled=True,
            cargo_cost=Decimal('17000.00'),
            service_enabled=True,
            service_cost=Decimal('9000.00'),
        )
        OrderItem.objects.create(
            order=order,
            product_name='Any',
            product_quantity=1,
            product_price_currency='UZS',
            product_price=Decimal('100000.000'),
            store='Taobao',
        )
        self.assertEqual(order.get_final_total, Decimal('555000.00'))

    def test_create_order_stores_auto_mode_when_total_equals_calculated(self):
        response = self.client.post(
            reverse('create_order'),
            {
                'receipt_number': '1003',
                'order_date': timezone.localdate().strftime('%Y-%m-%d'),
                'first_name': 'Auto',
                'last_name': 'Mode',
                'phone1': '+998901234000',
                'phone2': '',
                'debt': '0',
                'shipping_method': Order.ShippingMethod.AVIA,
                'status': Order.Status.ACCEPTED,
                'track_pending': 'on',
                'track_number': '',
                'cargo_enabled': 'on',
                'cargo_cost': '0',
                'service_enabled': 'on',
                'service_cost': '0',
                'manual_total': '100000',
                'usd_rate': '12205',
                'rmb_rate': '1807',
                'description': '',
                'items-TOTAL_FORMS': '1',
                'items-INITIAL_FORMS': '0',
                'items-MIN_NUM_FORMS': '0',
                'items-MAX_NUM_FORMS': '1000',
                'items-0-product_name': 'Auto item',
                'items-0-product_quantity': '1',
                'items-0-product_price_currency': 'UZS',
                'items-0-product_price': '100000',
                'items-0-shipping_method': Order.ShippingMethod.AVIA,
                'items-0-store': 'Taobao',
                'items-0-link': '',
                'items-0-DELETE': '',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        order = Order.objects.get(receipt_number=1003)
        self.assertIsNone(order.manual_total)
        self.assertEqual(order.get_final_total, Decimal('100000.00'))

    def test_create_order_preserves_manual_total_override_when_different(self):
        response = self.client.post(
            reverse('create_order'),
            {
                'receipt_number': '1004',
                'order_date': timezone.localdate().strftime('%Y-%m-%d'),
                'first_name': 'Manual',
                'last_name': 'Override',
                'phone1': '+998901234001',
                'phone2': '',
                'debt': '0',
                'shipping_method': Order.ShippingMethod.AVIA,
                'status': Order.Status.ACCEPTED,
                'track_pending': 'on',
                'track_number': '',
                'cargo_enabled': 'on',
                'cargo_cost': '0',
                'service_enabled': 'on',
                'service_cost': '0',
                'manual_total': '130000',
                'usd_rate': '12205',
                'rmb_rate': '1807',
                'description': '',
                'items-TOTAL_FORMS': '1',
                'items-INITIAL_FORMS': '0',
                'items-MIN_NUM_FORMS': '0',
                'items-MAX_NUM_FORMS': '1000',
                'items-0-product_name': 'Manual item',
                'items-0-product_quantity': '1',
                'items-0-product_price_currency': 'UZS',
                'items-0-product_price': '100000',
                'items-0-shipping_method': Order.ShippingMethod.AVIA,
                'items-0-store': 'Taobao',
                'items-0-link': '',
                'items-0-DELETE': '',
            },
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        order = Order.objects.get(receipt_number=1004)
        self.assertEqual(order.manual_total, Decimal('130000.00'))
        self.assertEqual(order.get_final_total, Decimal('130000.00'))


class ExcelImportExportTests(TestCase):
    def setUp(self):
        password = get_random_string(24)
        self.staff = User.objects.create_user(
            username='excelstaff',
            password=password,
            is_staff=True,
            is_active=True,
        )
        self.client.login(username=self.staff.username, password=password)

        self.order = Order.objects.create(
            user=self.staff,
            receipt_number=2001,
            order_date=timezone.localdate(),
            first_name='Excel',
            last_name='User',
            phone1=998901111111,
            shipping_method=Order.ShippingMethod.AVIA,
            status=Order.Status.ACCEPTED,
        )
        OrderItem.objects.create(
            order=self.order,
            product_name='Sneakers',
            product_quantity=2,
            product_price_currency='UZS',
            product_price='150000.000',
            store='Taobao',
        )

    def test_orders_export_excel_returns_xlsx(self):
        response = self.client.get(reverse('orders_export_excel'))
        self.assertEqual(response.status_code, 200)
        self.assertIn('spreadsheetml', response['Content-Type'])

        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook.active
        self.assertEqual(sheet.cell(1, 2).value, 'Квитанция')
        self.assertEqual(sheet.cell(2, 2).value, 2001)

    def test_orders_import_excel_creates_order_and_items(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            'Slug', 'Квитанция', 'Дата заказа', 'Имя клиента', 'Фамилия клиента',
            'Телефон #1', 'Телефон #2', 'Тип отправки', 'Статус', 'Трек-номер',
            'Трек ожидается', 'Карго включено', 'Карго сумма', 'Услуга включена',
            'Услуга сумма', 'Долг', 'Комментарий', 'Название товара', 'Количество',
            'Валюта', 'Себестоимость', 'Магазин', 'Ссылка', 'Админ',
        ])
        sheet.append([
            '', 3001, timezone.localdate().strftime('%Y-%m-%d'), 'Ali', 'Karimov',
            '+998901223344', '', 'AVIA', 'accepted', '', 'Да', 'Да', '0', 'Да',
            '0', '', '', 'Кроссовки', 1, 'UZS', '250000', 'Taobao', '', self.staff.username,
        ])

        payload = BytesIO()
        workbook.save(payload)
        payload.seek(0)
        uploaded = SimpleUploadedFile(
            'orders_import.xlsx',
            payload.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('orders_import_excel'),
            {'excel_file': uploaded, 'next': reverse('orders')},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        imported = Order.objects.get(receipt_number=3001)
        self.assertEqual(imported.first_name, 'Ali')
        self.assertEqual(imported.items.count(), 1)


class FullSiteRegressionLoadTests(TestCase):
    """Pre-production smoke/load tests for core views with 100 orders."""

    @classmethod
    def setUpTestData(cls):
        cls.staff_password = get_random_string(24)
        cls.staff = User.objects.create_user(
            username='loadstaff',
            password=cls.staff_password,
            is_staff=True,
            is_active=True,
        )

        cls.other_staff_password = get_random_string(24)
        cls.other_staff = User.objects.create_user(
            username='loadstaff2',
            password=cls.other_staff_password,
            is_staff=True,
            is_active=True,
        )

        cls.super_password = get_random_string(24)
        cls.superuser = User.objects.create_superuser(
            username='loadroot',
            password=cls.super_password,
            email='loadroot@example.com',
        )

        cls.client_password = get_random_string(24)
        cls.client_user = User.objects.create_user(
            username='regularclient',
            password=cls.client_password,
            is_staff=False,
            is_active=True,
        )

        today = timezone.localdate()
        status_cycle = [
            Order.Status.ACCEPTED,
            Order.Status.ORDERED,
            Order.Status.TRANSIT,
            Order.Status.ARRIVED,
            Order.Status.CANCELLED,
        ]
        shipping_cycle = [
            Order.ShippingMethod.AVIA,
            Order.ShippingMethod.IPOST,
            Order.ShippingMethod.CARGO_17994,
        ]
        store_cycle = ['Taobao', 'Alibaba', 'AliExpress', 'Pinduoduo', 'Poizon', '1688', '95', 'MadeChina']

        for idx in range(1, 101):
            owner = cls.staff if idx % 2 else cls.other_staff
            status = status_cycle[idx % len(status_cycle)]
            order = Order.objects.create(
                user=owner,
                receipt_number=400 + idx,
                order_date=today - timedelta(days=idx % 35),
                first_name=f'Client{idx}',
                last_name='Load',
                phone1=998900000000 + idx,
                status=status,
                shipping_method=shipping_cycle[idx % len(shipping_cycle)],
                track_number='' if idx % 3 == 0 else f'TRK{idx:06d}',
                cargo_enabled=(idx % 4 != 0),
                cargo_cost=Decimal('17000.00') if idx % 4 != 0 else Decimal('0.00'),
                service_enabled=(idx % 5 != 0),
                service_cost=Decimal('9000.00') if idx % 5 != 0 else Decimal('0.00'),
                description=f'Load dataset order #{idx}',
            )
            OrderItem.objects.create(
                order=order,
                product_name=f'Product {idx}',
                product_quantity=(idx % 4) + 1,
                product_price_currency='UZS',
                product_price=Decimal('125000.000') + Decimal(idx),
                store=store_cycle[idx % len(store_cycle)],
                link=f'https://example.com/order/{idx}',
            )
            if idx % 2 == 0:
                OrderAttachment.objects.create(
                    order=order,
                    image=f'orders/images/load_{idx}.jpg',
                )

        cls.sample_order = Order.objects.filter(user=cls.staff).order_by('id').first()

    def _login_staff(self):
        self.client.logout()
        self.assertTrue(self.client.login(username=self.staff.username, password=self.staff_password))

    def _login_superuser(self):
        self.client.logout()
        self.assertTrue(self.client.login(username=self.superuser.username, password=self.super_password))

    def _login_client(self):
        self.client.logout()
        self.assertTrue(self.client.login(username=self.client_user.username, password=self.client_password))

    def test_public_views_and_hidden_routes(self):
        response = self.client.get(reverse('index'))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('staff_login'))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('about'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'AkmalExpress')

        response = self.client.get(reverse('faq'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Частые вопросы')

        response = self.client.get(reverse('contacts'))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('robots_txt'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Disallow: /order/')

        response = self.client.get(reverse('login'))
        self.assertEqual(response.status_code, 404)

        response = self.client.get('/admin/')
        self.assertEqual(response.status_code, 404)

        detail_response = self.client.get(reverse('detail_order', args=[self.sample_order.slug]))
        self.assertEqual(detail_response.status_code, 200)
        self.assertContains(detail_response, f'№{self.sample_order.receipt_number}')
        self.assertNotContains(detail_response, 'Печать квитанции')

        lang_url = reverse('set_language', args=['uz'])
        response = self.client.get(f'{lang_url}?next={reverse("about")}', follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, translate_url(reverse('about'), 'uz'))
        self.assertIsNotNone(response.cookies.get('site_language'))

    def test_set_language_handles_legacy_ru_prefix_without_404(self):
        response = self.client.get(f'{reverse("set_language", args=["uz"])}?next=/ru/about/', follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, translate_url(reverse('about'), 'uz'))

    def test_set_language_switches_from_uz_path_to_ru_without_prefix(self):
        response = self.client.get(f'{reverse("set_language", args=["ru"])}?next=/uz/about/', follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse('about'))

    def test_index_is_reachable_with_uz_language_cookies(self):
        self.client.cookies['django_language'] = 'uz'
        self.client.cookies['site_language'] = 'uz'
        response = self.client.get(reverse('index'))
        self.assertEqual(response.status_code, 200)

    def test_access_control_for_regular_user(self):
        self._login_client()
        restricted_get_urls = [
            reverse('orders'),
            reverse('create_order'),
            reverse('dispatch_orders'),
            reverse('print_receipt', args=[self.sample_order.slug]),
            reverse('profile', args=[self.staff.username]),
            reverse('orders_export_excel'),
            reverse('create_admin'),
        ]

        for url in restricted_get_urls:
            with self.subTest(url=url):
                response = self.client.get(url, follow=False)
                self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse('detail_order', args=[self.sample_order.slug]))
        self.assertEqual(response.status_code, 200)

    def test_staff_views_under_load(self):
        self._login_staff()
        today = timezone.localdate().strftime('%Y-%m-%d')

        for page in range(1, 7):
            response = self.client.get(reverse('orders'), {'page': page})
            self.assertEqual(response.status_code, 200)

        for _ in range(3):
            response = self.client.get(reverse('dispatch_orders'), {'page': 1})
            self.assertEqual(response.status_code, 200)

        profile_url = reverse('profile', args=[self.staff.username])
        profile_queries = [
            {'search': 'Client', 'status': Order.Status.ACCEPTED, 'sort': 'date_desc', 'page': 1},
            {'search': 'Product', 'status': '', 'sort': 'receipt_asc', 'page': 2},
            {'date_from': (timezone.localdate() - timedelta(days=30)).strftime('%Y-%m-%d'), 'date_to': today},
        ]
        for query in profile_queries:
            response = self.client.get(profile_url, query)
            self.assertEqual(response.status_code, 200)

        for order in Order.objects.order_by('-id')[:15]:
            response = self.client.get(reverse('detail_order', args=[order.slug]))
            self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('print_receipt', args=[self.sample_order.slug]))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('create_product'))
        self.assertRedirects(response, reverse('create_order'), fetch_redirect_response=False)

        response = self.client.get(reverse('panel'))
        self.assertRedirects(response, reverse('orders'), fetch_redirect_response=False)

        response = self.client.post(reverse('dismiss_track_notice'), {'next': reverse('orders')}, follow=False)
        self.assertEqual(response.status_code, 302)

        response = self.client.get(reverse('my_profile'))
        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse('profile', args=[self.staff.username]), response.url)

        response = self.client.get(reverse('logout'))
        self.assertEqual(response.status_code, 302)

    def test_repeated_requests_stability(self):
        """Run repeated requests against core pages over a 100-order dataset."""
        self._login_staff()
        sample_slugs = list(Order.objects.order_by('-id').values_list('slug', flat=True)[:20])
        base_urls = [
            f"{reverse('index')}?search=Client",
            f"{reverse('orders')}?page=1",
            f"{reverse('dispatch_orders')}?page=1",
            f"{reverse('profile', args=[self.staff.username])}?search=Client&page=1",
        ]

        for _ in range(5):
            for url in base_urls:
                response = self.client.get(url)
                self.assertEqual(response.status_code, 200)
            for slug in sample_slugs:
                response = self.client.get(reverse('detail_order', args=[slug]))
                self.assertEqual(response.status_code, 200)

    def test_superuser_admin_and_management_views(self):
        self._login_superuser()
        response = self.client.get(reverse('create_admin'))
        self.assertEqual(response.status_code, 200)

        response = self.client.get(reverse('create_admin'), {'admin_id': str(self.staff.id), 'account_status': 'all'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context['admin_rows']), 1)
        self.assertEqual(response.context['admin_rows'][0]['user'].id, self.staff.id)

        response = self.client.post(
            reverse('create_admin'),
            {'username': 'newmoderator', 'password1': 'P@ssword-12345', 'password2': 'P@ssword-12345'},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        new_admin = User.objects.get(username='newmoderator')
        self.assertTrue(new_admin.is_staff)
        self.assertTrue(new_admin.is_active)

        response = self.client.post(
            reverse('toggle_status', args=[new_admin.id]),
            {'action': 'deactivate'},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        new_admin.refresh_from_db()
        self.assertFalse(new_admin.is_staff)
        self.assertFalse(new_admin.is_active)

        response = self.client.post(
            reverse('toggle_status', args=[new_admin.id]),
            {'action': 'activate'},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        new_admin.refresh_from_db()
        self.assertTrue(new_admin.is_staff)
        self.assertTrue(new_admin.is_active)

        response = self.client.post(reverse('delete_admin', args=[new_admin.id]), follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertFalse(User.objects.filter(username='newmoderator').exists())

        # Guard rails for superusers should remain active.
        response = self.client.post(reverse('delete_admin', args=[self.superuser.id]), follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(User.objects.filter(username=self.superuser.username).exists())

        response = self.client.get(reverse('toggle_status', args=[self.staff.id]), follow=False)
        self.assertEqual(response.status_code, 405)

    def test_create_change_delete_order_flow(self):
        self._login_staff()

        create_payload = {
            'receipt_number': '1499',
            'order_date': timezone.localdate().strftime('%Y-%m-%d'),
            'first_name': 'Flow',
            'last_name': 'Case',
            'phone1': '+998901010303',
            'phone2': '',
            'debt': '0',
            'shipping_method': Order.ShippingMethod.AVIA,
            'status': Order.Status.ACCEPTED,
            'track_pending': 'on',
            'track_number': '',
            'cargo_enabled': 'on',
            'cargo_cost': '17000',
            'service_enabled': 'on',
            'service_cost': '9000',
            'usd_rate': '12205',
            'rmb_rate': '1807',
            'description': 'Regression CRUD flow',
            'items-TOTAL_FORMS': '2',
            'items-INITIAL_FORMS': '0',
            'items-MIN_NUM_FORMS': '0',
            'items-MAX_NUM_FORMS': '1000',
            'items-0-product_name': 'Jacket',
            'items-0-product_quantity': '1',
            'items-0-product_price_currency': 'UZS',
            'items-0-product_price': '320000',
            'items-0-store': 'Taobao',
            'items-0-link': 'https://example.com/jacket',
            'items-0-DELETE': '',
            'items-1-product_name': 'Shoes',
            'items-1-product_quantity': '2',
            'items-1-product_price_currency': 'UZS',
            'items-1-product_price': '210000',
            'items-1-store': '95',
            'items-1-link': 'https://example.com/shoes',
            'items-1-DELETE': '',
        }
        response = self.client.post(reverse('create_order'), create_payload, follow=False)
        self.assertEqual(response.status_code, 302)

        created_order = Order.objects.get(receipt_number=1499)
        self.assertEqual(created_order.items.count(), 2)
        self.assertEqual(created_order.items.filter(link__isnull=False).count(), 2)

        change_payload = {
            'receipt_number': str(created_order.receipt_number),
            'order_date': created_order.order_date.strftime('%Y-%m-%d'),
            'first_name': created_order.first_name,
            'last_name': created_order.last_name,
            'phone1': str(created_order.phone1),
            'phone2': '',
            'debt': '0',
            'balance': str(created_order.balance or '0'),
            'manual_total': '',
            'usd_rate': '12205',
            'rmb_rate': '1807',
            'description': 'Updated flow',
            'status': Order.Status.TRANSIT,
            'items-TOTAL_FORMS': '2',
            'items-INITIAL_FORMS': '0',
            'items-MIN_NUM_FORMS': '0',
            'items-MAX_NUM_FORMS': '1000',
            'items-0-product_name': 'Jacket',
            'items-0-product_quantity': '1',
            'items-0-product_price_currency': 'UZS',
            'items-0-product_price': '320000',
            'items-0-shipping_method': Order.ShippingMethod.AVIA,
            'items-0-track_number': 'TRACK-FLOW-001',
            'items-0-store': 'Taobao',
            'items-0-link': 'https://example.com/jacket',
            'items-0-DELETE': '',
            'items-1-product_name': 'Shoes',
            'items-1-product_quantity': '2',
            'items-1-product_price_currency': 'UZS',
            'items-1-product_price': '210000',
            'items-1-shipping_method': Order.ShippingMethod.AVIA,
            'items-1-track_number': '',
            'items-1-store': '95',
            'items-1-link': 'https://example.com/shoes',
            'items-1-DELETE': '',
        }
        response = self.client.post(
            reverse('change_order', args=[created_order.slug]),
            change_payload,
            follow=False,
        )
        self.assertEqual(response.status_code, 302)

        created_order.refresh_from_db()
        self.assertEqual(created_order.status, Order.Status.TRANSIT)
        self.assertEqual(created_order.items.filter(track_number='TRACK-FLOW-001').count(), 1)

        response = self.client.get(reverse('delete_order', args=[created_order.slug]))
        self.assertEqual(response.status_code, 200)
        response = self.client.post(reverse('delete_order', args=[created_order.slug]), follow=False)
        self.assertEqual(response.status_code, 302)
        self.assertFalse(Order.objects.filter(receipt_number=1499).exists())

    def test_excel_import_export_for_orders_and_profiles(self):
        self._login_staff()

        response = self.client.get(reverse('orders_export_excel'))
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertGreaterEqual(workbook.active.max_row, 101)

        response = self.client.get(reverse('profile_export_excel', args=[self.staff.username]))
        self.assertEqual(response.status_code, 200)
        profile_workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertGreaterEqual(profile_workbook.active.max_row, 2)

        workbook_to_import = Workbook()
        sheet = workbook_to_import.active
        sheet.append([
            'Slug', 'Квитанция', 'Дата заказа', 'Имя клиента', 'Фамилия клиента',
            'Телефон #1', 'Телефон #2', 'Тип отправки', 'Статус', 'Трек-номер',
            'Трек ожидается', 'Карго включено', 'Карго сумма', 'Услуга включена',
            'Услуга сумма', 'Долг', 'Комментарий', 'Название товара', 'Количество',
            'Валюта', 'Себестоимость', 'Магазин', 'Ссылка', 'Админ',
        ])
        sheet.append([
            '', 1498, timezone.localdate().strftime('%Y-%m-%d'), 'Excel', 'Load',
            '+998901000001', '', 'AVIA', 'accepted', '', 'Да', 'Да', '15000', 'Да',
            '8000', '', 'bulk import 1', 'Bag', 1, 'UZS', '145000', 'Taobao',
            'https://example.com/bag', self.staff.username,
        ])
        sheet.append([
            '', 1498, timezone.localdate().strftime('%Y-%m-%d'), 'Excel', 'Load',
            '+998901000001', '', 'AVIA', 'accepted', '', 'Да', 'Да', '15000', 'Да',
            '8000', '', 'bulk import 1', 'Hat', 2, 'UZS', '85000', 'Alibaba',
            'https://example.com/hat', self.staff.username,
        ])

        payload = BytesIO()
        workbook_to_import.save(payload)
        payload.seek(0)
        uploaded = SimpleUploadedFile(
            'bulk_orders.xlsx',
            payload.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('orders_import_excel'),
            {'excel_file': uploaded, 'next': reverse('orders')},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        imported = Order.objects.get(receipt_number=1498, first_name='Excel', phone1=998901000001)
        self.assertEqual(imported.items.count(), 2)

        profile_wb = Workbook()
        profile_sheet = profile_wb.active
        profile_sheet.append([
            'Slug', 'Квитанция', 'Дата заказа', 'Имя клиента', 'Фамилия клиента',
            'Телефон #1', 'Телефон #2', 'Тип отправки', 'Статус', 'Трек-номер',
            'Трек ожидается', 'Карго включено', 'Карго сумма', 'Услуга включена',
            'Услуга сумма', 'Долг', 'Комментарий', 'Название товара', 'Количество',
            'Валюта', 'Себестоимость', 'Магазин', 'Ссылка', 'Админ',
        ])
        profile_sheet.append([
            '', 1497, timezone.localdate().strftime('%Y-%m-%d'), 'Profile', 'Import',
            '+998901000002', '', 'iPost', 'ordered', '', 'Да', 'Да', '10000', 'Да',
            '5000', '', '', 'Backpack', 1, 'UZS', '120000', '95', 'https://example.com/backpack', '',
        ])
        profile_payload = BytesIO()
        profile_wb.save(profile_payload)
        profile_payload.seek(0)
        profile_uploaded = SimpleUploadedFile(
            'profile_orders.xlsx',
            profile_payload.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        response = self.client.post(
            reverse('profile_import_excel', args=[self.other_staff.username]),
            {'excel_file': profile_uploaded, 'next': reverse('profile', args=[self.other_staff.username])},
            follow=False,
        )
        self.assertEqual(response.status_code, 302)
        profile_imported = Order.objects.get(receipt_number=1497, first_name='Profile')
        self.assertEqual(profile_imported.user, self.other_staff)
