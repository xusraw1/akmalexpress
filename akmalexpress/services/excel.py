from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.db import transaction
from django.db.models import Max, Min
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from akmalexpress.models import Order, OrderItem, Product

EXCEL_HEADERS = [
    'Slug',
    'Квитанция',
    'Дата заказа',
    'Имя клиента',
    'Фамилия клиента',
    'Телефон #1',
    'Телефон #2',
    'Тип отправки',
    'Статус',
    'Трек-номер',
    'Трек ожидается',
    'Курс USD',
    'Курс RMB',
    'Долг',
    'Комментарий',
    'Название товара',
    'Количество',
    'Валюта',
    'Себестоимость',
    'Магазин',
    'Ссылка',
    'Админ',
]

EXCEL_REQUIRED_KEYS = [
    'receipt_number',
    'order_date',
    'first_name',
    'phone1',
    'product_name',
    'product_quantity',
    'product_price_currency',
    'product_price',
    'store',
]

EXCEL_HEADER_ALIASES = {
    'slug': 'slug',
    'квитанция': 'receipt_number',
    'номер квитанции': 'receipt_number',
    'receipt': 'receipt_number',
    'receipt number': 'receipt_number',
    'дата заказа': 'order_date',
    'order date': 'order_date',
    'имя клиента': 'first_name',
    'first name': 'first_name',
    'фамилия клиента': 'last_name',
    'last name': 'last_name',
    'телефон #1': 'phone1',
    'телефон 1': 'phone1',
    'phone #1': 'phone1',
    'phone1': 'phone1',
    'телефон #2': 'phone2',
    'телефон 2': 'phone2',
    'phone #2': 'phone2',
    'phone2': 'phone2',
    'тип отправки': 'shipping_method',
    'shipping method': 'shipping_method',
    'статус': 'status',
    'status': 'status',
    'трек-номер': 'track_number',
    'трек номер': 'track_number',
    'track number': 'track_number',
    'track': 'track_number',
    'трек ожидается': 'track_pending',
    'track pending': 'track_pending',
    'курс usd': 'usd_rate',
    'usd rate': 'usd_rate',
    'курс rmb': 'rmb_rate',
    'курс cny': 'rmb_rate',
    'rmb rate': 'rmb_rate',
    'cny rate': 'rmb_rate',
    'долг': 'debt',
    'debt': 'debt',
    'комментарий': 'description',
    'description': 'description',
    'название товара': 'product_name',
    'product name': 'product_name',
    'количество': 'product_quantity',
    'qty': 'product_quantity',
    'quantity': 'product_quantity',
    'валюта': 'product_price_currency',
    'currency': 'product_price_currency',
    'себестоимость': 'product_price',
    'price': 'product_price',
    'cost': 'product_price',
    'магазин': 'store',
    'store': 'store',
    'ссылка': 'link',
    'link': 'link',
    'админ': 'admin_username',
    'admin': 'admin_username',
    'user': 'admin_username',
}

BOOL_TRUE_VALUES = {'1', 'true', 'yes', 'y', 'да', 'ha', 'on'}
BOOL_FALSE_VALUES = {'0', 'false', 'no', 'n', 'нет', "yo'q", 'off'}

STATUS_IMPORT_MAP = {
    'accepted': Order.Status.ACCEPTED,
    'принят': Order.Status.ACCEPTED,
    'qabul qilindi': Order.Status.ACCEPTED,
    'ordered': Order.Status.ORDERED,
    'заказан': Order.Status.ORDERED,
    'buyurtma qilindi': Order.Status.ORDERED,
    'transit': Order.Status.TRANSIT,
    'в пути': Order.Status.TRANSIT,
    "yo'lda": Order.Status.TRANSIT,
    'arrived': Order.Status.ARRIVED,
    'прибыл': Order.Status.ARRIVED,
    'yetib keldi': Order.Status.ARRIVED,
    'cancelled': Order.Status.CANCELLED,
    'отмена': Order.Status.CANCELLED,
    'bekor qilingan': Order.Status.CANCELLED,
}

SHIPPING_IMPORT_MAP = {
    'avia': Order.ShippingMethod.AVIA,
    'авиа': Order.ShippingMethod.AVIA,
    'ipost': Order.ShippingMethod.IPOST,
    '17994': Order.ShippingMethod.CARGO_17994,
}

STORE_IMPORT_MAP = {
    'none': Product.Store.NO,
    'taobao': Product.Store.TAOBAO,
    'alibaba': Product.Store.ALIBABA,
    'aliexpress': Product.Store.ALIEXPRESS,
    'pinduoduo': Product.Store.PINDUODUO,
    'poizon': Product.Store.POIZON,
    '1688': Product.Store.SIX,
    '95': Product.Store.NINETY_FIVE,
    'made in china': Product.Store.MADE_IN_CHINA,
    'madechina': Product.Store.MADE_IN_CHINA,
}

CURRENCY_IMPORT_MAP = {
    'usd': Product.Currency.USD,
    'uzs': Product.Currency.UZS,
    'rmb': Product.Currency.RMB,
    'cny': Product.Currency.RMB,
}

EXPORT_EXCEL_HEADERS = [
    'Дата заказа',
    'Квитанция',
    'Имя и фамилия',
    'Номер телефона',
    'Товары',
    'Себестоимость товаров',
    'Статус',
    'Долг',
    'Общая сумма',
]


def _safe_text(value):
    if value is None:
        return ''
    return str(value).strip()


def _normalize_header(value):
    return _safe_text(value).casefold().replace('№', '').replace('#', '').replace('  ', ' ')


def _resolve_excel_headers(header_row):
    resolved = {}
    for idx, raw_header in enumerate(header_row):
        alias = EXCEL_HEADER_ALIASES.get(_normalize_header(raw_header))
        if alias and alias not in resolved:
            resolved[alias] = idx
    return resolved


def _row_cell(row, headers, key):
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_excel_phone(value):
    digits = ''.join(ch for ch in _safe_text(value) if ch.isdigit())
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def _parse_excel_decimal(value, default=Decimal('0.00'), allow_none=False):
    text = _safe_text(value).replace(' ', '').replace(',', '.')
    if text == '':
        return None if allow_none else default

    try:
        return Decimal(text)
    except Exception:
        return default


def _parse_excel_int(value, default=0):
    text = _safe_text(value)
    if text == '':
        return default

    try:
        return int(float(text))
    except (ValueError, TypeError):
        return default


def _parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _safe_text(value)
    if not text:
        return None

    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_excel_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0

    text = _safe_text(value).casefold()
    if text in BOOL_TRUE_VALUES:
        return True
    if text in BOOL_FALSE_VALUES:
        return False
    return default


def _normalize_excel_status(value):
    key = _safe_text(value).casefold()
    return STATUS_IMPORT_MAP.get(key, Order.Status.ACCEPTED)


def _normalize_excel_shipping(value):
    key = _safe_text(value).casefold()
    return SHIPPING_IMPORT_MAP.get(key, Order.ShippingMethod.AVIA)


def _normalize_excel_store(value):
    key = _safe_text(value).casefold()
    return STORE_IMPORT_MAP.get(key, Product.Store.NO)


def _normalize_excel_currency(value):
    key = _safe_text(value).casefold()
    return CURRENCY_IMPORT_MAP.get(key, Product.Currency.UZS)


def _collect_order_items(order):
    items = list(order.items.all())
    if items:
        return [
            {
                'name': item.product_name,
                'quantity': item.product_quantity,
                'currency': item.product_price_currency,
                'price': item.product_price,
            }
            for item in items
        ]

    if order.product:
        return [
            {
                'name': order.product.product_name,
                'quantity': order.product.product_quantity,
                'currency': order.product.product_price_currency,
                'price': order.product.product_price,
            }
        ]
    return []


def _compose_order_datetime(order):
    if not order.order_date:
        return None
    if order.created_at:
        local_created_at = timezone.localtime(order.created_at)
        order_time = local_created_at.time().replace(microsecond=0)
    else:
        order_time = datetime.min.time()
    return datetime.combine(order.order_date, order_time)


def _format_decimal_for_text(value, precision=2):
    return f'{Decimal(value or "0.00"):.{precision}f}'


def _build_orders_workbook(orders_queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Orders'
    worksheet.append(EXPORT_EXCEL_HEADERS)

    header_font = Font(bold=True, color='132333')
    header_fill = PatternFill(fill_type='solid', fgColor='DCEFFB')
    zebra_fill_a = PatternFill(fill_type='solid', fgColor='FFFFFF')
    zebra_fill_b = PatternFill(fill_type='solid', fgColor='F6FBFF')
    wrap_alignment = Alignment(vertical='top', wrap_text=True)
    plain_alignment = Alignment(vertical='top')

    for col_idx, _ in enumerate(EXPORT_EXCEL_HEADERS, start=1):
        header_cell = worksheet.cell(row=1, column=col_idx)
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.alignment = Alignment(vertical='center', wrap_text=True)

    for order in orders_queryset:
        order_items = _collect_order_items(order)
        goods = '\n'.join(
            f'{item["name"]} × {item["quantity"]}'
            for item in order_items
        ) or '-'
        goods_costs = '\n'.join(
            f'{_format_decimal_for_text(item["price"], precision=3)} {item["currency"]}'
            for item in order_items
        ) or '-'

        worksheet.append(
            [
                _compose_order_datetime(order),
                order.receipt_number,
                f'{(order.first_name or "").strip()} {(order.last_name or "").strip()}'.strip(),
                str(order.phone1 or order.phone2 or ''),
                goods,
                goods_costs,
                order.get_status_display(),
                order.debt if order.debt is not None else Decimal('0.00'),
                order.get_final_total,
            ]
        )

    worksheet.freeze_panes = 'A2'

    for row_idx in range(2, worksheet.max_row + 1):
        row_fill = zebra_fill_a if row_idx % 2 == 0 else zebra_fill_b
        for col_idx in range(1, len(EXPORT_EXCEL_HEADERS) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            if col_idx in {5, 6}:
                cell.alignment = wrap_alignment
            else:
                cell.alignment = plain_alignment

    for row_idx in range(2, worksheet.max_row + 1):
        worksheet.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd hh:mm'
        worksheet.cell(row=row_idx, column=8).number_format = '#,##0.00 "UZS"'
        worksheet.cell(row=row_idx, column=9).number_format = '#,##0.00 "UZS"'

    for col_idx, title in enumerate(EXPORT_EXCEL_HEADERS, start=1):
        max_length = len(title)
        for row in worksheet.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            if isinstance(value, datetime):
                value = value.strftime('%Y-%m-%d %H:%M')
            raw_text = str(value)
            longest_line = max((len(line) for line in raw_text.splitlines()), default=0)
            max_length = max(max_length, longest_line)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 54)

    return workbook


def _excel_workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _build_export_filename(orders_queryset):
    period = orders_queryset.order_by().aggregate(period_start=Min('order_date'), period_end=Max('order_date'))
    period_start = period.get('period_start')
    period_end = period.get('period_end')
    if period_start and period_end:
        return f'orders_{period_start:%Y-%m-%d}_{period_end:%Y-%m-%d}.xlsx'
    return f'orders_{timezone.localtime().strftime("%Y%m%d_%H%M")}.xlsx'


def _import_orders_from_workbook(workbook, acting_user, fallback_user=None):
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError('Excel файл пустой.')

    headers = _resolve_excel_headers(header_row)
    missing_headers = [key for key in EXCEL_REQUIRED_KEYS if key not in headers]
    if missing_headers:
        missing_list = ', '.join(missing_headers)
        raise ValueError(f'Не найдены обязательные колонки: {missing_list}.')

    grouped_orders = {}
    skipped_rows = 0
    row_errors = []

    for row_number, row in enumerate(rows, start=2):
        if not any(_safe_text(cell) for cell in row):
            continue

        try:
            receipt_number = _parse_excel_int(_row_cell(row, headers, 'receipt_number'))
            if receipt_number < 1:
                raise ValueError('некорректный номер квитанции')

            order_date = _parse_excel_date(_row_cell(row, headers, 'order_date'))
            if not order_date:
                raise ValueError('некорректная дата заказа')

            first_name = _safe_text(_row_cell(row, headers, 'first_name'))
            if not first_name:
                raise ValueError('не указано имя клиента')

            last_name = _safe_text(_row_cell(row, headers, 'last_name'))
            phone1 = _parse_excel_phone(_row_cell(row, headers, 'phone1'))
            if phone1 is None:
                raise ValueError('некорректный телефон #1')

            slug = _safe_text(_row_cell(row, headers, 'slug'))
            key = f'slug:{slug}' if slug else f'{receipt_number}|{order_date.isoformat()}|{first_name}|{last_name}|{phone1}'

            track_number = _safe_text(_row_cell(row, headers, 'track_number'))
            track_pending = _parse_excel_bool(_row_cell(row, headers, 'track_pending'), default=not bool(track_number))
            if track_pending:
                track_number = ''

            usd_rate = _parse_excel_decimal(_row_cell(row, headers, 'usd_rate'), default=Decimal('12205.00'))
            rmb_rate = _parse_excel_decimal(_row_cell(row, headers, 'rmb_rate'), default=Decimal('1807.00'))

            order_payload = {
                'slug': slug,
                'receipt_number': receipt_number,
                'order_date': order_date,
                'first_name': first_name,
                'last_name': last_name,
                'phone1': phone1,
                'phone2': _parse_excel_phone(_row_cell(row, headers, 'phone2')),
                'shipping_method': _normalize_excel_shipping(_row_cell(row, headers, 'shipping_method')),
                'status': _normalize_excel_status(_row_cell(row, headers, 'status')),
                'track_number': track_number,
                'usd_rate': usd_rate,
                'rmb_rate': rmb_rate,
                'debt': _parse_excel_decimal(_row_cell(row, headers, 'debt'), allow_none=True),
                'description': _safe_text(_row_cell(row, headers, 'description')),
                'admin_username': _safe_text(_row_cell(row, headers, 'admin_username')),
            }

            grouped = grouped_orders.setdefault(key, {'order': order_payload, 'items': []})
            if order_payload['admin_username'] and not grouped['order'].get('admin_username'):
                grouped['order']['admin_username'] = order_payload['admin_username']

            product_name = _safe_text(_row_cell(row, headers, 'product_name'))
            if product_name:
                quantity = max(1, _parse_excel_int(_row_cell(row, headers, 'product_quantity'), default=1))
                grouped['items'].append(
                    {
                        'product_name': product_name[:140],
                        'product_quantity': quantity,
                        'product_price_currency': _normalize_excel_currency(_row_cell(row, headers, 'product_price_currency')),
                        'product_price': _parse_excel_decimal(_row_cell(row, headers, 'product_price')),
                        'track_number': track_number,
                        'store': _normalize_excel_store(_row_cell(row, headers, 'store')),
                        'link': _safe_text(_row_cell(row, headers, 'link')) or None,
                    }
                )

        except ValueError as exc:
            skipped_rows += 1
            if len(row_errors) < 8:
                row_errors.append(f'Строка {row_number}: {exc}')

    if not grouped_orders:
        raise ValueError('В Excel не найдены строки с заказами.')

    created_orders = 0
    updated_orders = 0
    imported_items = 0

    with transaction.atomic():
        for payload in grouped_orders.values():
            order_data = payload['order']
            items_data = payload['items']

            assigned_user = fallback_user or acting_user
            admin_username = order_data.get('admin_username', '').lstrip('@')
            if admin_username:
                admin_user = User.objects.filter(username__iexact=admin_username, is_active=True).first()
                if admin_user and (admin_user.is_staff or admin_user.is_superuser):
                    assigned_user = admin_user

            order = None
            if order_data['slug']:
                order = Order.objects.filter(slug=order_data['slug']).first()

            if order is None:
                order = Order.objects.filter(
                    receipt_number=order_data['receipt_number'],
                    order_date=order_data['order_date'],
                    first_name=order_data['first_name'],
                    last_name=order_data['last_name'],
                    phone1=order_data['phone1'],
                ).first()

            is_new = order is None
            if is_new:
                order = Order()
                if order_data['slug']:
                    order.slug = order_data['slug']
                created_orders += 1
            else:
                updated_orders += 1

            order.user = assigned_user
            order.product = None
            order.receipt_number = order_data['receipt_number']
            order.order_date = order_data['order_date']
            order.shipping_method = order_data['shipping_method']
            order.track_number = order_data['track_number']
            order.first_name = order_data['first_name']
            order.last_name = order_data['last_name']
            order.phone1 = order_data['phone1']
            order.phone2 = order_data['phone2']
            order.debt = order_data['debt']
            order.cargo_enabled = False
            order.cargo_cost = Decimal('0.00')
            order.service_enabled = False
            order.service_cost = Decimal('0.00')
            order.usd_rate = order_data['usd_rate']
            order.rmb_rate = order_data['rmb_rate']
            order.description = order_data['description']
            order.status = order_data['status']
            order.come = timezone.now() if order.status == Order.Status.ARRIVED else None
            order.save()

            order.items.all().delete()
            for item_data in items_data:
                OrderItem.objects.create(order=order, **item_data)
                imported_items += 1

    return {
        'created_orders': created_orders,
        'updated_orders': updated_orders,
        'imported_items': imported_items,
        'skipped_rows': skipped_rows,
        'row_errors': row_errors,
    }
