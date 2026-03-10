from collections import defaultdict
from calendar import monthrange
from datetime import date, datetime, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO
from urllib.parse import parse_qsl, urlencode, urlparse, urlunparse

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.models import User
from django.conf import settings
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Q
from django.http import HttpResponse, HttpResponseNotAllowed, HttpResponseNotFound
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils import timezone
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from .context_processors import TRACK_NOTICE_DISMISS_KEY
from .forms import ChangeOrderForm, CreateOrderForm, OrderItemFormSet, save_order_items
from .i18n import normalize_language
from .models import Order, OrderItem, Product


def orders_with_related(queryset):
    """Attach related entities to reduce N+1 queries across order pages."""
    return queryset.select_related('product', 'user').prefetch_related('items', 'attachments')


def parse_month_filter(value):
    """Parse YYYY-MM from query params."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m")
    except ValueError:
        return None


def parse_date_filter(value):
    """Parse YYYY-MM-DD from query params."""
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def get_period_bounds(period, anchor_date):
    """Return inclusive date range for day/week/month filters."""
    if period == 'week':
        start_date = anchor_date - timedelta(days=anchor_date.weekday())
        end_date = start_date + timedelta(days=6)
        title = 'За неделю'
    elif period == 'month':
        start_date = anchor_date.replace(day=1)
        end_day = monthrange(anchor_date.year, anchor_date.month)[1]
        end_date = anchor_date.replace(day=end_day)
        title = 'За месяц'
    else:
        start_date = anchor_date
        end_date = anchor_date
        title = 'За день'
    return start_date, end_date, title


def configure_order_item_formset(item_formset):
    """Hide DELETE controls; rows are managed by custom UI buttons."""
    for form in item_formset.forms:
        if 'DELETE' in form.fields:
            form.fields['DELETE'].widget.attrs.update(
                {
                    'hidden': 'hidden',
                    'tabindex': '-1',
                    'aria-hidden': 'true',
                    'class': 'row-delete-input',
                }
            )
    return item_formset


EXCEL_HEADERS = [
    'Slug',
    'Квитанция',
    'Дата заказа',
    'Имя клиента',
    'Фамилия клиента',
    'Телефон #1',
    'Телефон #2',
    'Тип отправки',
    'Статус',
    'Трек-номер',
    'Трек ожидается',
    'Карго включено',
    'Карго сумма',
    'Услуга включена',
    'Услуга сумма',
    'Курс USD',
    'Курс RMB',
    'Долг',
    'Комментарий',
    'Название товара',
    'Количество',
    'Валюта',
    'Себестоимость',
    'Магазин',
    'Ссылка',
    'Админ',
]

EXCEL_REQUIRED_KEYS = [
    'receipt_number',
    'order_date',
    'first_name',
    'phone1',
    'product_name',
    'product_quantity',
    'product_price_currency',
    'product_price',
    'store',
]

EXCEL_HEADER_ALIASES = {
    'slug': 'slug',
    'квитанция': 'receipt_number',
    'номер квитанции': 'receipt_number',
    'receipt': 'receipt_number',
    'receipt number': 'receipt_number',
    'дата заказа': 'order_date',
    'order date': 'order_date',
    'имя клиента': 'first_name',
    'first name': 'first_name',
    'фамилия клиента': 'last_name',
    'last name': 'last_name',
    'телефон #1': 'phone1',
    'телефон 1': 'phone1',
    'phone #1': 'phone1',
    'phone1': 'phone1',
    'телефон #2': 'phone2',
    'телефон 2': 'phone2',
    'phone #2': 'phone2',
    'phone2': 'phone2',
    'тип отправки': 'shipping_method',
    'shipping method': 'shipping_method',
    'статус': 'status',
    'status': 'status',
    'трек-номер': 'track_number',
    'трек номер': 'track_number',
    'track number': 'track_number',
    'track': 'track_number',
    'трек ожидается': 'track_pending',
    'track pending': 'track_pending',
    'карго включено': 'cargo_enabled',
    'cargo enabled': 'cargo_enabled',
    'карго сумма': 'cargo_cost',
    'cargo cost': 'cargo_cost',
    'услуга включена': 'service_enabled',
    'service enabled': 'service_enabled',
    'услуга сумма': 'service_cost',
    'service cost': 'service_cost',
    'курс usd': 'usd_rate',
    'usd rate': 'usd_rate',
    'курс rmb': 'rmb_rate',
    'курс cny': 'rmb_rate',
    'rmb rate': 'rmb_rate',
    'cny rate': 'rmb_rate',
    'долг': 'debt',
    'debt': 'debt',
    'комментарий': 'description',
    'description': 'description',
    'название товара': 'product_name',
    'product name': 'product_name',
    'количество': 'product_quantity',
    'qty': 'product_quantity',
    'quantity': 'product_quantity',
    'валюта': 'product_price_currency',
    'currency': 'product_price_currency',
    'себестоимость': 'product_price',
    'price': 'product_price',
    'cost': 'product_price',
    'магазин': 'store',
    'store': 'store',
    'ссылка': 'link',
    'link': 'link',
    'админ': 'admin_username',
    'admin': 'admin_username',
    'user': 'admin_username',
}

BOOL_TRUE_VALUES = {'1', 'true', 'yes', 'y', 'да', 'ha', 'on'}
BOOL_FALSE_VALUES = {'0', 'false', 'no', 'n', 'нет', "yo'q", 'off'}

STATUS_IMPORT_MAP = {
    'accepted': Order.Status.ACCEPTED,
    'принят': Order.Status.ACCEPTED,
    'qabul qilindi': Order.Status.ACCEPTED,
    'ordered': Order.Status.ORDERED,
    'заказан': Order.Status.ORDERED,
    'buyurtma qilindi': Order.Status.ORDERED,
    'transit': Order.Status.TRANSIT,
    'в пути': Order.Status.TRANSIT,
    "yo'lda": Order.Status.TRANSIT,
    'arrived': Order.Status.ARRIVED,
    'прибыл': Order.Status.ARRIVED,
    'yetib keldi': Order.Status.ARRIVED,
    'cancelled': Order.Status.CANCELLED,
    'отмена': Order.Status.CANCELLED,
    'bekor qilingan': Order.Status.CANCELLED,
}

SHIPPING_IMPORT_MAP = {
    'avia': Order.ShippingMethod.AVIA,
    'авиа': Order.ShippingMethod.AVIA,
    'ipost': Order.ShippingMethod.IPOST,
    '17994': Order.ShippingMethod.CARGO_17994,
}

STORE_IMPORT_MAP = {
    'none': Product.Store.NO,
    'taobao': Product.Store.TAOBAO,
    'alibaba': Product.Store.ALIBABA,
    'aliexpress': Product.Store.ALIEXPRESS,
    'pinduoduo': Product.Store.PINDUODUO,
    'poizon': Product.Store.POIZON,
    '1688': Product.Store.SIX,
    '95': Product.Store.NINETY_FIVE,
    'made in china': Product.Store.MADE_IN_CHINA,
    'madechina': Product.Store.MADE_IN_CHINA,
}

CURRENCY_IMPORT_MAP = {
    'usd': Product.Currency.USD,
    'uzs': Product.Currency.UZS,
    'rmb': Product.Currency.RMB,
    'cny': Product.Currency.RMB,
}


def _safe_text(value):
    if value is None:
        return ''
    return str(value).strip()


def _normalize_header(value):
    return _safe_text(value).casefold().replace('№', '').replace('#', '').replace('  ', ' ')


def _resolve_excel_headers(header_row):
    resolved = {}
    for idx, raw_header in enumerate(header_row):
        alias = EXCEL_HEADER_ALIASES.get(_normalize_header(raw_header))
        if alias and alias not in resolved:
            resolved[alias] = idx
    return resolved


def _row_cell(row, headers, key):
    idx = headers.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_excel_phone(value):
    digits = ''.join(ch for ch in _safe_text(value) if ch.isdigit())
    if not digits:
        return None
    try:
        return int(digits)
    except ValueError:
        return None


def _parse_excel_decimal(value, default=Decimal('0.00'), allow_none=False):
    text = _safe_text(value).replace(' ', '').replace(',', '.')
    if text == '':
        return None if allow_none else default

    try:
        return Decimal(text)
    except Exception:
        return default


def _parse_excel_int(value, default=0):
    text = _safe_text(value)
    if text == '':
        return default

    try:
        return int(float(text))
    except (ValueError, TypeError):
        return default


def _parse_excel_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = _safe_text(value)
    if not text:
        return None

    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_excel_bool(value, default=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0

    text = _safe_text(value).casefold()
    if text in BOOL_TRUE_VALUES:
        return True
    if text in BOOL_FALSE_VALUES:
        return False
    return default


def _normalize_excel_status(value):
    key = _safe_text(value).casefold()
    return STATUS_IMPORT_MAP.get(key, Order.Status.ACCEPTED)


def _normalize_excel_shipping(value):
    key = _safe_text(value).casefold()
    return SHIPPING_IMPORT_MAP.get(key, Order.ShippingMethod.AVIA)


def _normalize_excel_store(value):
    key = _safe_text(value).casefold()
    return STORE_IMPORT_MAP.get(key, Product.Store.NO)


def _normalize_excel_currency(value):
    key = _safe_text(value).casefold()
    return CURRENCY_IMPORT_MAP.get(key, Product.Currency.UZS)


def _safe_next_redirect(request, fallback_url):
    next_url = request.POST.get('next') or request.GET.get('next') or request.META.get('HTTP_REFERER') or fallback_url
    if not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return fallback_url
    return next_url


def _build_orders_workbook(orders_queryset):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Orders'
    worksheet.append(EXCEL_HEADERS)

    for order in orders_queryset:
        items = list(order.items.all())
        if items:
            items_rows = [
                {
                    'name': item.product_name,
                    'quantity': item.product_quantity,
                    'currency': item.product_price_currency,
                    'price': item.product_price,
                    'store': item.store,
                    'link': item.link or '',
                }
                for item in items
            ]
        elif order.product:
            items_rows = [{
                'name': order.product.product_name,
                'quantity': order.product.product_quantity,
                'currency': order.product.product_price_currency,
                'price': order.product.product_price,
                'store': order.product.store,
                'link': order.product.link or '',
            }]
        else:
            items_rows = [{
                'name': '',
                'quantity': 1,
                'currency': order.get_current or Product.Currency.UZS,
                'price': Decimal('0.00'),
                'store': Product.Store.NO,
                'link': '',
            }]

        for item_data in items_rows:
            worksheet.append(
                [
                    order.slug or '',
                    order.receipt_number,
                    order.order_date.strftime('%Y-%m-%d') if order.order_date else '',
                    order.first_name or '',
                    order.last_name or '',
                    str(order.phone1 or ''),
                    str(order.phone2 or ''),
                    order.shipping_method,
                    order.status,
                    order.track_number or '',
                    'Да' if not (order.track_number or '').strip() else 'Нет',
                    'Да' if order.cargo_enabled else 'Нет',
                    str(order.cargo_cost or Decimal('0.00')),
                    'Да' if order.service_enabled else 'Нет',
                    str(order.service_cost or Decimal('0.00')),
                    str(order.usd_rate or Decimal('12205.00')),
                    str(order.rmb_rate or Decimal('1807.00')),
                    '' if order.debt is None else str(order.debt),
                    order.description or '',
                    item_data['name'],
                    item_data['quantity'],
                    item_data['currency'],
                    str(item_data['price']),
                    item_data['store'],
                    item_data['link'],
                    order.user.username if order.user else '',
                ]
            )

    worksheet.freeze_panes = 'A2'
    for col_idx, title in enumerate(EXCEL_HEADERS, start=1):
        max_length = len(title)
        for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 42)

    return workbook


def _excel_workbook_response(workbook, filename):
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _import_orders_from_workbook(workbook, acting_user, fallback_user=None):
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        raise ValueError('Excel файл пустой.')

    headers = _resolve_excel_headers(header_row)
    missing_headers = [key for key in EXCEL_REQUIRED_KEYS if key not in headers]
    if missing_headers:
        missing_list = ', '.join(missing_headers)
        raise ValueError(f'Не найдены обязательные колонки: {missing_list}.')

    grouped_orders = {}
    skipped_rows = 0
    row_errors = []

    for row_number, row in enumerate(rows, start=2):
        if not any(_safe_text(cell) for cell in row):
            continue

        try:
            receipt_number = _parse_excel_int(_row_cell(row, headers, 'receipt_number'))
            if receipt_number < 1:
                raise ValueError('некорректный номер квитанции')

            order_date = _parse_excel_date(_row_cell(row, headers, 'order_date'))
            if not order_date:
                raise ValueError('некорректная дата заказа')

            first_name = _safe_text(_row_cell(row, headers, 'first_name'))
            if not first_name:
                raise ValueError('не указано имя клиента')

            last_name = _safe_text(_row_cell(row, headers, 'last_name'))
            phone1 = _parse_excel_phone(_row_cell(row, headers, 'phone1'))
            if phone1 is None:
                raise ValueError('некорректный телефон #1')

            slug = _safe_text(_row_cell(row, headers, 'slug'))
            key = f'slug:{slug}' if slug else f'{receipt_number}|{order_date.isoformat()}|{first_name}|{last_name}|{phone1}'

            track_number = _safe_text(_row_cell(row, headers, 'track_number'))
            track_pending = _parse_excel_bool(_row_cell(row, headers, 'track_pending'), default=not bool(track_number))
            if track_pending:
                track_number = ''

            cargo_enabled = _parse_excel_bool(_row_cell(row, headers, 'cargo_enabled'), default=True)
            service_enabled = _parse_excel_bool(_row_cell(row, headers, 'service_enabled'), default=True)
            cargo_cost = _parse_excel_decimal(_row_cell(row, headers, 'cargo_cost'))
            service_cost = _parse_excel_decimal(_row_cell(row, headers, 'service_cost'))
            usd_rate = _parse_excel_decimal(_row_cell(row, headers, 'usd_rate'), default=Decimal('12205.00'))
            rmb_rate = _parse_excel_decimal(_row_cell(row, headers, 'rmb_rate'), default=Decimal('1807.00'))
            if not cargo_enabled:
                cargo_cost = Decimal('0.00')
            if not service_enabled:
                service_cost = Decimal('0.00')

            order_payload = {
                'slug': slug,
                'receipt_number': receipt_number,
                'order_date': order_date,
                'first_name': first_name,
                'last_name': last_name,
                'phone1': phone1,
                'phone2': _parse_excel_phone(_row_cell(row, headers, 'phone2')),
                'shipping_method': _normalize_excel_shipping(_row_cell(row, headers, 'shipping_method')),
                'status': _normalize_excel_status(_row_cell(row, headers, 'status')),
                'track_number': track_number,
                'cargo_enabled': cargo_enabled,
                'cargo_cost': cargo_cost,
                'service_enabled': service_enabled,
                'service_cost': service_cost,
                'usd_rate': usd_rate,
                'rmb_rate': rmb_rate,
                'debt': _parse_excel_decimal(_row_cell(row, headers, 'debt'), allow_none=True),
                'description': _safe_text(_row_cell(row, headers, 'description')),
                'admin_username': _safe_text(_row_cell(row, headers, 'admin_username')),
            }

            grouped = grouped_orders.setdefault(key, {'order': order_payload, 'items': []})
            if order_payload['admin_username'] and not grouped['order'].get('admin_username'):
                grouped['order']['admin_username'] = order_payload['admin_username']

            product_name = _safe_text(_row_cell(row, headers, 'product_name'))
            if product_name:
                quantity = max(1, _parse_excel_int(_row_cell(row, headers, 'product_quantity'), default=1))
                grouped['items'].append(
                    {
                        'product_name': product_name[:140],
                        'product_quantity': quantity,
                        'product_price_currency': _normalize_excel_currency(_row_cell(row, headers, 'product_price_currency')),
                        'product_price': _parse_excel_decimal(_row_cell(row, headers, 'product_price')),
                        'store': _normalize_excel_store(_row_cell(row, headers, 'store')),
                        'link': _safe_text(_row_cell(row, headers, 'link')) or None,
                    }
                )

        except ValueError as exc:
            skipped_rows += 1
            if len(row_errors) < 8:
                row_errors.append(f'Строка {row_number}: {exc}')

    if not grouped_orders:
        raise ValueError('В Excel не найдены строки с заказами.')

    created_orders = 0
    updated_orders = 0
    imported_items = 0

    for payload in grouped_orders.values():
        order_data = payload['order']
        items_data = payload['items']

        assigned_user = fallback_user or acting_user
        admin_username = order_data.get('admin_username', '').lstrip('@')
        if admin_username:
            admin_user = User.objects.filter(username__iexact=admin_username, is_active=True).first()
            if admin_user and (admin_user.is_staff or admin_user.is_superuser):
                assigned_user = admin_user

        order = None
        if order_data['slug']:
            order = Order.objects.filter(slug=order_data['slug']).first()

        if order is None:
            order = Order.objects.filter(
                receipt_number=order_data['receipt_number'],
                order_date=order_data['order_date'],
                first_name=order_data['first_name'],
                last_name=order_data['last_name'],
                phone1=order_data['phone1'],
            ).first()

        is_new = order is None
        if is_new:
            order = Order()
            if order_data['slug']:
                order.slug = order_data['slug']
            created_orders += 1
        else:
            updated_orders += 1

        order.user = assigned_user
        order.product = None
        order.receipt_number = order_data['receipt_number']
        order.order_date = order_data['order_date']
        order.shipping_method = order_data['shipping_method']
        order.track_number = order_data['track_number']
        order.first_name = order_data['first_name']
        order.last_name = order_data['last_name']
        order.phone1 = order_data['phone1']
        order.phone2 = order_data['phone2']
        order.debt = order_data['debt']
        order.cargo_enabled = order_data['cargo_enabled']
        order.cargo_cost = order_data['cargo_cost']
        order.service_enabled = order_data['service_enabled']
        order.service_cost = order_data['service_cost']
        order.usd_rate = order_data['usd_rate']
        order.rmb_rate = order_data['rmb_rate']
        order.description = order_data['description']
        order.status = order_data['status']
        order.come = timezone.now() if order.status == Order.Status.ARRIVED else None
        order.save()

        order.items.all().delete()
        for item_data in items_data:
            OrderItem.objects.create(order=order, **item_data)
            imported_items += 1

    return {
        'created_orders': created_orders,
        'updated_orders': updated_orders,
        'imported_items': imported_items,
        'skipped_rows': skipped_rows,
        'row_errors': row_errors,
    }


def is_active_superuser(user):
    return user.is_staff or user.is_superuser


def user_is_order_creator(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        slug = kwargs.get('slug')
        order = Order.objects.filter(slug=slug).first()
        if not order:
            return redirect('/')

        if request.user.is_superuser or request.user.is_staff or request.user == order.user:
            return view_func(request, *args, **kwargs)

        messages.error(request, "У вас нет прав для доступа к этой странице")
        return redirect('/')

    return _wrapped_view


def superuser_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_superuser:
            messages.error(request, "У вас нет прав для доступа к этой странице")
            return redirect('/')
        return view_func(request, *args, **kwargs)

    return _wrapped_view


def index(request):
    """Global order search page with optional date range filter."""
    context = {}
    search = (request.GET.get('search') or '').strip()
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()
    date_from = parse_date_filter(date_from_raw) if date_from_raw else None
    date_to = parse_date_filter(date_to_raw) if date_to_raw else None

    if date_from_raw and not date_from:
        messages.warning(request, 'Неверный формат даты. Используйте YYYY-MM-DD.')
    if date_to_raw and not date_to:
        messages.warning(request, 'Неверный формат даты. Используйте YYYY-MM-DD.')

    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    context['search_query'] = search
    context['date_from'] = date_from.strftime('%Y-%m-%d') if date_from else ''
    context['date_to'] = date_to.strftime('%Y-%m-%d') if date_to else ''

    has_filters = bool(search or date_from or date_to)
    if has_filters:
        queryset = Order.objects.all()

        search_filter = (
            Q(receipt_number__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(items__product_name__icontains=search)
            | Q(product__product_name__icontains=search)
        )
        if search:
            if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
                search_filter |= Q(track_number__icontains=search)
            queryset = queryset.filter(search_filter)

        if date_from:
            queryset = queryset.filter(order_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(order_date__lte=date_to)

        orders_qs = orders_with_related(
            queryset
            .distinct()
            .order_by('-order_date', '-created_at')
        )

        if orders_qs.exists():
            paginator = Paginator(orders_qs, 5)
            page_number = request.GET.get('page')

            try:
                orders = paginator.page(page_number)
            except PageNotAnInteger:
                orders = paginator.page(1)
            except EmptyPage:
                orders = paginator.page(paginator.num_pages)

            if search and not page_number:
                messages.success(request, f"Заказы по вашему запросу '{search}' найдены")
            context['orders'] = orders
        else:
            if search:
                messages.info(request, f"По вашему запросу '{search}' ничего не найдено")
            else:
                messages.info(request, 'По вашим фильтрам ничего не найдено')

    return render(request, 'index.html', context)


def contacts_view(request):
    return render(request, 'akmalexpress/contacts.html')


def about_view(request):
    """Public company information page."""
    return render(request, 'akmalexpress/about.html')


def hidden_entrypoint(request):
    """Return 404 for legacy public URLs that should stay hidden."""
    return HttpResponseNotFound('Not found')


def panel_entrypoint(request):
    """Unified convenience entrypoint for staff/superusers."""
    if not request.user.is_authenticated:
        return redirect('staff_login')
    if request.user.is_superuser:
        return redirect(f"/{settings.ADMIN_URL}")
    if request.user.is_staff:
        return redirect('orders')
    return redirect('index')


def custom_404(request, exception):
    """Render branded 404 page for unknown routes."""
    return render(request, '404.html', status=404)


def custom_404_debug(request, unmatched_path=''):
    """Render 404 page for debug mode catch-all route."""
    return render(request, '404.html', status=404)


@user_passes_test(is_active_superuser)
def dismiss_track_notice(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    dismissed_until = timezone.now() + timedelta(days=2)
    request.session[TRACK_NOTICE_DISMISS_KEY] = dismissed_until.isoformat()

    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or '/'
    if not url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        next_url = '/'

    messages.info(request, 'Напоминание о трек-номерах скрыто на 2 дня.')
    return redirect(next_url)


def robots_txt(request):
    """Block indexing of private/admin pages for search crawlers."""
    admin_path = f"/{settings.ADMIN_URL}".replace('//', '/')
    staff_login_path = f"/{settings.STAFF_LOGIN_URL}".replace('//', '/')
    private_paths = list(dict.fromkeys([
        admin_path,
        admin_path.rstrip('/'),
        staff_login_path,
        staff_login_path.rstrip('/'),
        '/login/',
        '/logout/',
        '/panel/',
        '/panel',
        '/notifications/',
        '/order/',
        '/create/',
        '/toggle_status/',
        '/delete_admin/',
        '/profile/',
    ]))
    lines = ['User-agent: *']
    lines.extend(f'Disallow: {path}' for path in private_paths)
    lines.append('')
    return HttpResponse('\n'.join(lines), content_type='text/plain')


def set_language_view(request, lang_code):
    language = normalize_language(lang_code)
    request.session['site_language'] = language

    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER') or '/'
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}, require_https=request.is_secure()):
        next_url = '/'
    else:
        parsed_next = urlparse(next_url)
        clean_query = [(key, value) for key, value in parse_qsl(parsed_next.query, keep_blank_values=True) if key != 'lang']
        next_url = urlunparse(parsed_next._replace(query=urlencode(clean_query, doseq=True)))

    response = redirect(next_url)
    response.set_cookie('site_language', language, max_age=60 * 60 * 24 * 365)
    return response


def detail_order(request, slug):
    order = get_object_or_404(orders_with_related(Order.objects.all()), slug=slug)
    if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
        return render(request, 'akmalexpress/detail_order.html', {'order': order})
    return render(request, 'akmalexpress/client_order_detail.html', {'order': order})


@user_passes_test(is_active_superuser)
def print_receipt(request, slug):
    order = get_object_or_404(orders_with_related(Order.objects.all()), slug=slug)
    return render(request, 'akmalexpress/receipt_print.html', {'order': order})


@user_passes_test(is_active_superuser)
@user_is_order_creator
def delete_order(request, slug):
    order = get_object_or_404(Order, slug=slug)
    if request.method == 'POST':
        order.delete()
        messages.success(request, f"Заказ с номером №{order.receipt_number} успешно удален")
        return redirect('/')
    return render(request, 'akmalexpress/delete_order.html', {'order': order})


@user_passes_test(is_active_superuser)
@user_is_order_creator
def change_order(request, slug):
    orderr = get_object_or_404(Order, slug=slug)
    form = ChangeOrderForm(instance=orderr)

    if request.method == 'POST':
        form = ChangeOrderForm(request.POST, instance=orderr)

        if form.is_valid():
            order = form.save(commit=False)
            if order.status == Order.Status.ARRIVED:
                order.come = timezone.now()
            elif order.status != Order.Status.ARRIVED:
                order.come = None

            order.save()
            messages.success(request, f"Заказ с квитанцией №{order.receipt_number} обновлен")
            return redirect('orders')

        messages.warning(request, 'Введенные данные неверны')

    return render(request, 'akmalexpress/change_order.html', {'form': form, 'orderr': orderr})


@user_passes_test(is_active_superuser)
def create_order(request):
    last_order = Order.objects.order_by('-receipt_number').first()
    previous_receipt_number = last_order.receipt_number if last_order is not None else None
    next_receipt_number = (last_order.receipt_number + 1) if last_order is not None else 1

    form = CreateOrderForm(initial={'receipt_number': next_receipt_number})
    item_formset = configure_order_item_formset(OrderItemFormSet(prefix='items'))

    if request.method == 'POST':
        form = CreateOrderForm(request.POST, request.FILES)
        item_formset = configure_order_item_formset(OrderItemFormSet(request.POST, prefix='items'))

        if form.is_valid() and item_formset.is_valid():
            order = form.save_order(user=request.user)
            save_order_items(order, item_formset)

            messages.success(request, f'Заказ №{order.receipt_number} успешно создан')

            if not order.track_number:
                remind_at = order.order_date + timedelta(days=2)
                messages.warning(
                    request,
                    f"Трек-номер пока не добавлен. Проверьте заказ и добавьте трек до {remind_at.strftime('%d.%m.%Y')}.",
                )

            return redirect('orders')

        messages.warning(request, 'Проверьте форму: есть ошибки в данных заказа или товаров.')

    return render(
        request,
        'akmalexpress/create_order.html',
        {
            'form': form,
            'item_formset': item_formset,
            'receipt_number': previous_receipt_number,
            'next_receipt_number': next_receipt_number,
        },
    )


@user_passes_test(is_active_superuser)
def create_product(request):
    messages.info(request, 'Создание товара перенесено в единую форму создания заказа.')
    return redirect('create_order')


@user_passes_test(is_active_superuser)
def order_list(request):
    orders_list = Order.objects.all()
    search_query = (request.GET.get('search') or '').strip()

    if search_query:
        search_filter = (
            Q(receipt_number__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(phone1__icontains=search_query)
            | Q(phone2__icontains=search_query)
            | Q(items__product_name__icontains=search_query)
            | Q(product__product_name__icontains=search_query)
        )
        if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
            search_filter |= Q(track_number__icontains=search_query)
        orders_list = orders_list.filter(search_filter)

    selected_month = request.GET.get('month', '').strip()
    month_date = None
    if request.user.is_superuser and selected_month:
        month_date = parse_month_filter(selected_month)
        if month_date:
            orders_list = orders_list.filter(order_date__year=month_date.year, order_date__month=month_date.month)
        else:
            selected_month = ''
            messages.warning(request, 'Неверный формат месяца. Используйте YYYY-MM.')

    orders_list = orders_with_related(
        orders_list
        .distinct()
        .order_by('-order_date', '-created_at')
    )

    paginator = Paginator(orders_list, 20)
    page_number = request.GET.get('page')

    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    monthly_total = None
    monthly_orders_count = None
    monthly_admin_stats = []
    month_label = None

    if request.user.is_superuser:
        summary_month = month_date
        if summary_month is None:
            today = timezone.localdate()
            summary_month = datetime(year=today.year, month=today.month, day=1)

        month_label = summary_month.strftime('%m.%Y')
        monthly_scope = list(
            orders_with_related(
                Order.objects.filter(
                    order_date__year=summary_month.year,
                    order_date__month=summary_month.month,
                )
            )
        )
        monthly_total = sum((o.get_final_total for o in monthly_scope), Decimal('0.00'))
        monthly_orders_count = len(monthly_scope)

        grouped = defaultdict(lambda: {'orders_count': 0, 'total': Decimal('0.00')})
        for order in monthly_scope:
            username = order.user.username if order.user else '-'
            grouped[username]['orders_count'] += 1
            grouped[username]['total'] += order.get_final_total

        monthly_admin_stats = [
            {
                'user__username': username,
                'orders_count': data['orders_count'],
                'total': data['total'],
            }
            for username, data in grouped.items()
        ]
        monthly_admin_stats.sort(key=lambda x: x['total'], reverse=True)

    context = {
        'orders': orders,
        'search_query': search_query,
        'selected_month': selected_month,
        'monthly_total': monthly_total,
        'monthly_orders_count': monthly_orders_count,
        'monthly_admin_stats': monthly_admin_stats,
        'month_label': month_label,
    }
    return render(request, 'akmalexpress/orders.html', context)


@user_passes_test(is_active_superuser)
def export_orders_excel(request):
    orders_qs = Order.objects.all()
    search_query = (request.GET.get('search') or '').strip()

    if search_query:
        search_filter = (
            Q(receipt_number__icontains=search_query)
            | Q(first_name__icontains=search_query)
            | Q(last_name__icontains=search_query)
            | Q(phone1__icontains=search_query)
            | Q(phone2__icontains=search_query)
            | Q(items__product_name__icontains=search_query)
            | Q(product__product_name__icontains=search_query)
        )
        if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
            search_filter |= Q(track_number__icontains=search_query)
        orders_qs = orders_qs.filter(search_filter)

    selected_month = (request.GET.get('month') or '').strip()
    if request.user.is_superuser and selected_month:
        month_date = parse_month_filter(selected_month)
        if month_date:
            orders_qs = orders_qs.filter(order_date__year=month_date.year, order_date__month=month_date.month)

    orders_qs = orders_with_related(
        orders_qs
        .distinct()
        .order_by('-order_date', '-created_at')
    )
    workbook = _build_orders_workbook(orders_qs)
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return _excel_workbook_response(workbook, f'akmalexpress_orders_{stamp}.xlsx')


@user_passes_test(is_active_superuser)
def import_orders_excel(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    fallback_url = reverse('orders')
    next_url = _safe_next_redirect(request, fallback_url)
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.warning(request, 'Файл Excel не загружен.')
        return redirect(next_url)

    if not excel_file.name.lower().endswith('.xlsx'):
        messages.warning(request, 'Поддерживается только формат .xlsx.')
        return redirect(next_url)

    try:
        workbook = load_workbook(excel_file, data_only=True)
        result = _import_orders_from_workbook(workbook, acting_user=request.user, fallback_user=request.user)
    except ValueError as exc:
        messages.error(request, f'Ошибка импорта Excel: {exc}')
        return redirect(next_url)
    except Exception:
        messages.error(request, 'Не удалось прочитать Excel файл.')
        return redirect(next_url)

    messages.success(
        request,
        f"Excel импорт завершен: создано заказов {result['created_orders']}, обновлено {result['updated_orders']}, добавлено товаров {result['imported_items']}.",
    )
    if result['skipped_rows']:
        messages.warning(request, f"Пропущено строк: {result['skipped_rows']}.")
    for row_error in result['row_errors']:
        messages.warning(request, row_error)
    return redirect(next_url)


@user_passes_test(is_active_superuser)
def dispatch_orders_view(request):
    """Dispatch board: show only new accepted orders waiting to be ordered."""

    if request.method == 'POST':
        order_id = (request.POST.get('order_id') or '').strip()
        new_status = (request.POST.get('status') or '').strip()
        available_statuses = {choice[0] for choice in Order.Status.choices}

        if order_id.isdigit() and new_status in available_statuses:
            order = get_object_or_404(Order, id=int(order_id))
            order.status = new_status
            if new_status == Order.Status.ARRIVED:
                order.come = timezone.now()
            else:
                order.come = None
            order.save(update_fields=['status', 'come', 'updated_at'])
            messages.success(request, f'Статус заказа №{order.receipt_number} обновлен.')
        else:
            messages.warning(request, 'Не удалось обновить статус заказа.')

        return redirect('dispatch_orders')

    dispatch_orders_qs = orders_with_related(
        Order.objects.filter(
            status=Order.Status.ACCEPTED,
        ).order_by('-order_date', '-created_at')
    )

    total_orders = dispatch_orders_qs.count()
    total_items = sum(len(order.items.all()) for order in dispatch_orders_qs)

    paginator = Paginator(dispatch_orders_qs, 20)
    page_number = request.GET.get('page')
    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    context = {
        'orders': orders,
        'status_choices': Order.Status.choices,
        'total_orders': total_orders,
        'total_items': total_items,
    }
    return render(request, 'akmalexpress/dispatch_orders.html', context)


@user_passes_test(is_active_superuser)
def my_profile_redirect(request):
    target_url = reverse('profile', kwargs={'user': request.user.username})
    query = request.GET.urlencode()
    if query:
        target_url = f'{target_url}?{query}'
    return redirect(target_url)


def _resolve_profile_user(user):
    normalized_user = (user or '').strip().lstrip('@')
    if not normalized_user:
        return None
    return User.objects.filter(username__iexact=normalized_user).first()


@user_passes_test(is_active_superuser)
def profile_view(request, user):
    """Admin profile page with filtering/sorting over user's orders."""
    profile = _resolve_profile_user(user)
    if profile is None:
        messages.error(request, 'Пользователь не найден')
        return redirect('index')

    if not (profile.is_superuser or profile.is_staff or profile.is_active):
        messages.warning(request, 'Вы не имеете доступ к этому профилю')
        return redirect('index')

    status_choices = list(Order.Status.choices)
    allowed_statuses = {choice[0] for choice in status_choices}
    allowed_sorts = {
        'date_desc': ('-order_date', '-created_at'),
        'date_asc': ('order_date', 'created_at'),
        'receipt_desc': ('-receipt_number', '-created_at'),
        'receipt_asc': ('receipt_number', 'created_at'),
        'status_asc': ('status', '-order_date'),
        'status_desc': ('-status', '-order_date'),
    }

    search = (request.GET.get('search') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    date_from_raw = (request.GET.get('date_from') or '').strip()
    date_to_raw = (request.GET.get('date_to') or '').strip()
    sort = (request.GET.get('sort') or 'date_desc').strip()

    date_from = parse_date_filter(date_from_raw) if date_from_raw else None
    date_to = parse_date_filter(date_to_raw) if date_to_raw else None

    if date_from_raw and not date_from:
        messages.warning(request, 'Неверный формат даты. Используйте YYYY-MM-DD.')
    if date_to_raw and not date_to:
        messages.warning(request, 'Неверный формат даты. Используйте YYYY-MM-DD.')
    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from

    if status_filter and status_filter not in allowed_statuses:
        status_filter = ''

    if sort not in allowed_sorts:
        sort = 'date_desc'

    all_profile_orders_qs = orders_with_related(
        Order.objects.filter(user=profile).order_by('-order_date', '-created_at')
    )
    profile_total_amount = sum((o.get_final_total for o in all_profile_orders_qs), Decimal('0.00'))
    orders_total_count = all_profile_orders_qs.count()

    profile_orders_qs = Order.objects.filter(user=profile)
    if search:
        search_filter = (
            Q(receipt_number__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(items__product_name__icontains=search)
            | Q(product__product_name__icontains=search)
        )
        if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
            search_filter |= Q(track_number__icontains=search)
        profile_orders_qs = profile_orders_qs.filter(search_filter)

    if status_filter:
        profile_orders_qs = profile_orders_qs.filter(status=status_filter)
    if date_from:
        profile_orders_qs = profile_orders_qs.filter(order_date__gte=date_from)
    if date_to:
        profile_orders_qs = profile_orders_qs.filter(order_date__lte=date_to)

    profile_orders_qs = orders_with_related(
        profile_orders_qs
        .distinct()
        .order_by(*allowed_sorts[sort])
    )
    filtered_total_amount = sum((o.get_final_total for o in profile_orders_qs), Decimal('0.00'))

    paginator = Paginator(profile_orders_qs, 10)
    page_number = request.GET.get('page')

    try:
        orders = paginator.page(page_number)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)

    reminder_date = timezone.localdate() - timedelta(days=2)
    profile_track_notice_qs = orders_with_related(
        Order.objects.filter(user=profile)
        .filter(
            order_date__lte=reminder_date,
            status__in=[Order.Status.ACCEPTED, Order.Status.ORDERED, Order.Status.TRANSIT],
        )
        .filter(Q(track_number__isnull=True) | Q(track_number=''))
        .order_by('order_date', '-created_at')
    )
    profile_track_notice_count = profile_track_notice_qs.count()
    notice_paginator = Paginator(profile_track_notice_qs, 5)
    notice_page_number = request.GET.get('notice_page')
    try:
        profile_track_notice_page = notice_paginator.page(notice_page_number)
    except PageNotAnInteger:
        profile_track_notice_page = notice_paginator.page(1) if profile_track_notice_count else None
    except EmptyPage:
        profile_track_notice_page = notice_paginator.page(notice_paginator.num_pages) if profile_track_notice_count else None

    return render(
        request,
        'akmalexpress/profile.html',
        {
            'profile': profile,
            'orders': orders,
            'profile_total_amount': profile_total_amount,
            'filtered_total_amount': filtered_total_amount,
            'orders_total_count': orders_total_count,
            'profile_track_notice_count': profile_track_notice_count,
            'profile_track_notice_page': profile_track_notice_page,
            'hide_global_track_banner': True,
            'status_choices': status_choices,
            'filters': {
                'search': search,
                'status': status_filter,
                'date_from': date_from.strftime('%Y-%m-%d') if date_from else '',
                'date_to': date_to.strftime('%Y-%m-%d') if date_to else '',
                'sort': sort,
            },
        },
    )


@user_passes_test(is_active_superuser)
def export_profile_orders_excel(request, user):
    profile = _resolve_profile_user(user)
    if profile is None:
        messages.error(request, 'Пользователь не найден')
        return redirect('index')

    status_choices = list(Order.Status.choices)
    allowed_statuses = {choice[0] for choice in status_choices}
    allowed_sorts = {
        'date_desc': ('-order_date', '-created_at'),
        'date_asc': ('order_date', 'created_at'),
        'receipt_desc': ('-receipt_number', '-created_at'),
        'receipt_asc': ('receipt_number', 'created_at'),
        'status_asc': ('status', '-order_date'),
        'status_desc': ('-status', '-order_date'),
    }

    search = (request.GET.get('search') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    date_from = parse_date_filter((request.GET.get('date_from') or '').strip())
    date_to = parse_date_filter((request.GET.get('date_to') or '').strip())
    sort = (request.GET.get('sort') or 'date_desc').strip()

    if date_from and date_to and date_from > date_to:
        date_from, date_to = date_to, date_from
    if status_filter and status_filter not in allowed_statuses:
        status_filter = ''
    if sort not in allowed_sorts:
        sort = 'date_desc'

    orders_qs = Order.objects.filter(user=profile)
    if search:
        search_filter = (
            Q(receipt_number__icontains=search)
            | Q(first_name__icontains=search)
            | Q(last_name__icontains=search)
            | Q(items__product_name__icontains=search)
            | Q(product__product_name__icontains=search)
        )
        if request.user.is_authenticated and (request.user.is_staff or request.user.is_superuser):
            search_filter |= Q(track_number__icontains=search)
        orders_qs = orders_qs.filter(search_filter)

    if status_filter:
        orders_qs = orders_qs.filter(status=status_filter)
    if date_from:
        orders_qs = orders_qs.filter(order_date__gte=date_from)
    if date_to:
        orders_qs = orders_qs.filter(order_date__lte=date_to)

    orders_qs = orders_with_related(orders_qs.distinct().order_by(*allowed_sorts[sort]))
    workbook = _build_orders_workbook(orders_qs)
    stamp = timezone.localtime().strftime('%Y%m%d_%H%M')
    return _excel_workbook_response(workbook, f'akmalexpress_orders_{profile.username}_{stamp}.xlsx')


@user_passes_test(is_active_superuser)
def import_profile_orders_excel(request, user):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    profile = _resolve_profile_user(user)
    if profile is None:
        messages.error(request, 'Пользователь не найден')
        return redirect('index')

    fallback_url = reverse('profile', kwargs={'user': profile.username})
    next_url = _safe_next_redirect(request, fallback_url)
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        messages.warning(request, 'Файл Excel не загружен.')
        return redirect(next_url)

    if not excel_file.name.lower().endswith('.xlsx'):
        messages.warning(request, 'Поддерживается только формат .xlsx.')
        return redirect(next_url)

    try:
        workbook = load_workbook(excel_file, data_only=True)
        result = _import_orders_from_workbook(workbook, acting_user=request.user, fallback_user=profile)
    except ValueError as exc:
        messages.error(request, f'Ошибка импорта Excel: {exc}')
        return redirect(next_url)
    except Exception:
        messages.error(request, 'Не удалось прочитать Excel файл.')
        return redirect(next_url)

    messages.success(
        request,
        f"Excel импорт завершен: создано заказов {result['created_orders']}, обновлено {result['updated_orders']}, добавлено товаров {result['imported_items']}.",
    )
    if result['skipped_rows']:
        messages.warning(request, f"Пропущено строк: {result['skipped_rows']}.")
    for row_error in result['row_errors']:
        messages.warning(request, row_error)
    return redirect(next_url)


def login_view(request):
    if request.user.is_authenticated:
        messages.error(request, "Ошибка")
        return redirect('/')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if not (user.is_staff or user.is_superuser):
                messages.error(request, 'Служебный вход доступен только администраторам.')
                return redirect('staff_login')
            login(request, user)
            messages.success(request, 'Вы вошли в свой аккаунт')
            if user.is_staff or user.is_superuser:
                reminder_date = timezone.localdate() - timedelta(days=2)
                if user.is_superuser:
                    pending_qs = Order.objects.filter(
                        order_date__lte=reminder_date,
                        status__in=[Order.Status.ACCEPTED, Order.Status.ORDERED, Order.Status.TRANSIT],
                    )
                else:
                    pending_qs = Order.objects.filter(
                        user=user,
                        order_date__lte=reminder_date,
                        status__in=[Order.Status.ACCEPTED, Order.Status.ORDERED, Order.Status.TRANSIT],
                    )
                pending_count = pending_qs.filter(
                    Q(track_number__isnull=True) | Q(track_number='')
                ).count()
                if pending_count:
                    messages.info(
                        request,
                        f'У вас {pending_count} заказ(ов) без трек-номера старше 2 дней. Откройте список заказов и обновите их.',
                    )
            return redirect('/')
        messages.error(request, 'Пользователь не найден, попробуйте заново')
        return redirect('staff_login')

    return render(request, 'akmalexpress/login.html')


def logout_view(request):
    logout(request)
    messages.warning(request, "Вы вышли из аккаунта")
    return redirect('index')


@superuser_required
def toggle_status(request, user_id):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    user = get_object_or_404(User, id=user_id)
    if user.is_superuser:
        messages.warning(request, 'Суперпользователя нельзя деактивировать через эту форму.')
        return redirect('create_admin')

    action = request.POST.get('action')
    if action == 'activate':
        user.is_active = True
        user.is_staff = True
        user.save(update_fields=['is_active', 'is_staff'])
        messages.success(request, f"Модератор {user} активирован")
    elif action == 'deactivate':
        user.is_active = False
        user.is_staff = False
        user.save(update_fields=['is_active', 'is_staff'])
        messages.info(request, f"Модератор {user} деактивирован")

    return redirect('create_admin')


@superuser_required
def delete_admin(request, user_id):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])

    user = get_object_or_404(User, id=user_id)
    if user.is_superuser:
        messages.warning(request, 'Суперпользователя удалять через эту форму нельзя.')
        return redirect('create_admin')

    if user == request.user:
        messages.warning(request, 'Нельзя удалить текущего пользователя.')
        return redirect('create_admin')

    username = user.username
    user.delete()
    messages.success(request, f'Админ @{username} удален')
    return redirect('create_admin')


@superuser_required
def create_admin(request):
    users = User.objects.exclude(is_superuser=True).order_by('username')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password1 = request.POST.get('password1')
        password2 = request.POST.get('password2')
        user = User.objects.filter(username=username)

        if not user:
            if username and password1 == password2:
                created_user = User.objects.create_user(username=username, password=password1)
                created_user.is_staff = True
                created_user.is_active = True
                created_user.save(update_fields=['is_staff', 'is_active'])
                messages.success(request, f"Модератор @{username} успешно добавлен")
                return redirect('create_admin')

            messages.warning(request, "Пароли не совпадают или не указано имя пользователя")
            return redirect('create_admin')

        messages.warning(request, f"Пользователь с ником {username} уже существует!")
        return redirect('create_admin')

    return render(request, 'akmalexpress/create_admin.html', {'users': users})
